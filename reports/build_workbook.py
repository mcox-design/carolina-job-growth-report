"""Build the Carolinas Job Growth & Housing Demand running database (.xlsx).

One row per company/project announcement. Tabs:
Weekly Summary · Ranked Opportunities · Running Database · Markets to Watch ·
Excluded/Noise · Source Log · Scoring Methodology

Re-runnable: regenerates the workbook from the WEEK_* data structures below.
For future weeks, add a new WEEK_N_* block below the existing ones, concatenate
it into RUNNING / WATCH / EXCLUDED / SCORE_DETAIL, and bump REPORT_DATE/WINDOW.
Prior weeks' rows are never edited or discarded — only appended to.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

REPORT_DATE = "2026-07-13"
WINDOW = "2026-01-13 to 2026-07-13"  # trailing 6 months

PRIOR_REPORT_DATE = "2026-07-06"
PRIOR_WINDOW = "2026-01-06 to 2026-07-06"

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SUB_FONT = Font(italic=True, size=9, color="595959")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# ============================================================================
# Running Database: one row per announcement (full schema)
# ============================================================================
RUNNING_HEADERS = [
    "report_date", "status_wow", "rank", "score", "market_submarket", "county", "state",
    "company_project", "industry", "announcement_type", "job_count", "job_count_type",
    "salary_wage_stated", "salary_inferred", "capital_investment", "incentives", "timeline",
    "source_date", "source_url", "confidence", "job_type", "urban_suburban_rural",
    "housing_demand_implication", "notes",
]

# ---- Week 1 (2026-06-24, inaugural / baseline report) ----
WEEK1_RUNNING = [
    ["2026-06-24", "New", 1, 92, "Charlotte / Uptown", "Mecklenburg", "NC",
     "SMBC Group — 2nd US HQ / bank operations", "Banking / financial services",
     "New office / HQ-tier ops expansion", 2000, "Projected (6 yr)",
     "$165,316 avg (stated; vs county $90,706)", "N/A — stated", "$50.5M",
     "JDIG up to $70.0M / 12 yr; $23.3M IDF-Utility Account", "Jobs phased over 6 yr",
     "2026-04-07", "https://www.commerce.nc.gov/news/press-releases/2026/04/07/governor-stein-announces-2000-new-jobs-japans-smbc-group-selects-charlotte-bank-operations-expansion",
     "High", "Office / HQ banking", "Urban",
     "Premier high-income signal; Class-A urban MF + high-end for-sale TH in Uptown/South End/Dilworth.",
     "Avg pay ~1.8x county avg. Strongest deal of the quarter."],

    ["2026-06-24", "New (carry-over project, in-window milestone)", 2, 85, "Blythewood / N. Columbia", "Richland", "SC",
     "Scout Motors — EV assembly plant (hiring ramp)", "Automotive / EV manufacturing",
     "Hiring ramp + training-center milestone", 4000, "Projected (600+ hired)",
     "$30-$37.50/hr line (stated, news); salaried higher", "Estimated/Inferred mix supports above-median HH income",
     "$2.0B (+$25M training center)", "State/local package (2023); $25M readySC", "Production targeted end-2026; 4,000 at full capacity ~2030-31",
     "2026-04-20", "https://www.wistv.com/2026/04/20/scout-motors-hold-grand-opening-blythewood-area-training-center/",
     "High (project/milestone); Med (wage)", "Auto mfg + salaried eng.", "Suburban",
     "Bifurcated: salaried/professional -> higher-income for-sale & Class-A; hourly base -> workforce/missing-middle rental in N. Richland/Fairfield.",
     "Qualifying basis = in-window hiring milestone; original selection 2023."],

    ["2026-06-24", "New", 3, 84, "Durham / RTP", "Durham", "NC",
     "AbbVie — biopharma manufacturing campus", "Biopharma / life sciences mfg",
     "New facility (greenfield, 185 ac)", 734, "Projected (+2,000 construction)",
     "$118,041 avg (stated; vs county $102,817)", "N/A — stated", "$1.4B",
     "JDIG up to $19.3M / 12 yr; $6.4M IDF-Utility; +city/county", "Construction 2026; complete ~end-2028; hiring through ~2031",
     "2026-04-22", "https://news.abbvie.com/2026-04-22-AbbVie-Selects-North-Carolina-for-New-1-4-Billion-Manufacturing-Campus",
     "High", "Adv. mfg / R&D / lab", "Suburban-urban",
     "High-end MF + move-up for-sale around Durham/RTP; spillover Wake/Orange. Construction-phase rental near-term.",
     "With AbbVie, NC hosts 8 of 10 largest pharma firms by revenue."],

    ["2026-06-24", "New (carry-over project, in-window milestone)", 4, 82, "Greensboro / PTI Airport", "Guilford", "NC",
     "JetZero — aerospace plant (groundbreaking)", "Aerospace / advanced manufacturing",
     "Groundbreaking of megaproject", 14500, "Projected",
     "$89,340 avg (stated, 2025 selection)", "N/A — stated (2025)", "$4.7B",
     "JDIG up to ~$1.02B / 37 yr; Guilford grant ~$75.9M / 20 yr", "Groundbreaking 2026-06-15; buildout from 2026",
     "2026-06-15", "https://www.commerce.nc.gov/news/press-releases/2026/06/15/governor-stein-celebrates-jetzero-groundbreaking-launch-greensboro-airplane-makers-14500-job-project",
     "High (facts); Med (timeline)", "Aerospace mfg + eng.", "Suburban",
     "Transformational long-term: multi-yr construction then 14,500 jobs; broad rental + for-sale across Guilford + Randolph/Alamance commuter shed.",
     "Largest job commitment in NC history. CAUTION: reported state-budget hiring-timeline delay. Wage = 2025 figure."],

    ["2026-06-24", "New (carry-over project, in-window expansion)", 5, 80, "Holly Springs / SW Wake", "Wake", "NC",
     "Genentech — biomanufacturing expansion (investment doubled to ~$2B)", "Biopharma / life sciences mfg",
     "Expansion of committed project", 500, "Projected (site total; +100 in-window, +1,500 construction)",
     "~$119,833 avg (stated, original 2025 tranche; ~$120k, 56% above Wake avg)", "N/A — stated", "~$2.0B (doubled from ~$700M)",
     "JDIG up to $9.85M / 12 yr (original 420-job award; no new incentive on the +100)", "Operational by 2029",
     "2026-01-20", "https://www.gene.com/media/press-releases/15096/2026-01-20/genentech-more-than-doubles-investment-i",
     "High", "Adv. biomanufacturing", "Suburban",
     "Strong higher-income for-sale + Class-A rental in fast-growing SW Wake (Holly Springs/Apex/Fuquay); ~$120k wage well above county median. Near-term 1,500+ construction workforce.",
     "In-window event = Jan 2026 investment doubling that pushes site over 500 jobs; the new increment is +100. Wage figure traces to original 2025 award."],

    ["2026-06-24", "New", 6, 78, "Charlotte", "Mecklenburg", "NC",
     "Capital Group — East Coast operations hub", "Investment mgmt / finance + tech",
     "New facility / operations hub", 600, "Projected",
     "Not disclosed", "Estimated/Inferred ~$190k (from ~$116M payroll / 600 jobs; roles all $100k+ in CLT)", "$60M",
     "JDIG up to $17.2M / 12 yr; $5.7M IDF-Utility", "12-yr JDIG term",
     "2026-03-26", "https://governor.nc.gov/news/press-releases/2026/03/26/governor-stein-announces-capital-group-will-establish-major-operations-hub-charlotte",
     "High (jobs); Med (wage)", "Office — finance/tech", "Urban-suburban",
     "High-skill eng/data workforce -> walkable South End/Uptown & inner-suburban Class-A and for-sale TH.",
     "Wage not officially stated; ~$190k is derived (payroll math), labeled Inferred."],
]

# ---- Week 2 (2026-07-06) ----
WEEK2_RUNNING = [
    ["2026-07-06", "Gaining Momentum", 1, 93, "Charlotte / Uptown", "Mecklenburg", "NC",
     "SMBC Group — 2nd US HQ / bank operations", "Banking / financial services",
     "Lease signed / active hiring milestone", 2000, "Projected (6 yr)",
     "$165,316 avg (stated; vs county $90,706)", "N/A — stated", "$50.5M",
     "JDIG up to $70.0M / 12 yr; $23.3M IDF-Utility Account", "Leased former One Wells Fargo Center; hiring 70+ roles; first workers fall 2026",
     "2026-05-04", "https://www.axios.com/local/charlotte/2026/05/04/smbc-headquarters-wells-fargo-center-uptown",
     "High", "Office / HQ banking", "Urban",
     "Premier high-income signal; Class-A urban MF + high-end for-sale TH in Uptown/South End/Dilworth.",
     "Real-estate + hiring milestone confirms April announcement is executing."],

    ["2026-07-06", "New", 2, 88, "Rock Hill / York Co.", "York", "SC",
     "Octapharma — \"Project Palmetto Rock\" HQ + manufacturing campus", "Biopharmaceutical (plasma fractionation)",
     "New US HQ + manufacturing campus", 1500, "Projected (1,200 new + 300 relocated)",
     "$141,502 avg HQ / $102,752 avg mfg (both stated, county-presented figures)", "N/A — stated", "$1.5B",
     "Fee-in-lieu-of-tax, assessment ratio cut to 4%; City of Rock Hill forgoes its share; Rock Hill Schools share cut to 31.9%",
     "2nd reading approved 2026-06-29 (5-2); construction could start late 2026, build-out ~10 yrs",
     "2026-06-29", "https://www.postandcourier.com/york-county/news/rock-hill-plasma-panthers-octapharma-billion/article_fbd570b1-6451-406b-bb8c-92a53fad0d7c.html",
     "High (facts); Med (deal certainty)", "Office/HQ + biopharma mfg", "Suburban",
     "Dual-tier: $141k HQ roles -> luxury for-sale/Class-A in Rock Hill/Fort Mill; $102k mfg roles -> premium workforce-plus rental/TH.",
     "First SC deal in this dataset with two stated six-figure wage tiers. Tax-share amendment is a real execution risk flagged by local press."],

    ["2026-07-06", "Gaining Momentum", 3, 86, "Blythewood / N. Columbia", "Richland", "SC",
     "Scout Motors — EV assembly plant (production milestone)", "Automotive / EV manufacturing",
     "First vehicle body welded / production ramp", 4000, "Projected (600+ hired)",
     "$30-$37.50/hr line (stated, news); salaried higher", "Estimated/Inferred mix supports above-median HH income",
     "$2.0B (+$25M training center)", "State/local package (2023); $25M readySC", "First Traveler body welded June 2026; production targeted end-2026",
     "2026-06-01", "https://blog.scoutmotors.com/june-2026-scout-motors-production-center-update/",
     "High", "Auto mfg + salaried eng.", "Suburban",
     "Bifurcated: salaried/professional -> higher-income for-sale & Class-A; hourly base -> workforce/missing-middle rental in N. Richland/Fairfield.",
     "First-ever vehicle-body weld is a materially lower-risk milestone than training-center opening alone."],

    ["2026-07-06", "Repeated", 4, 84, "Durham / RTP", "Durham", "NC",
     "AbbVie — biopharma manufacturing campus", "Biopharma / life sciences mfg",
     "New facility (greenfield, 185 ac)", 734, "Projected (+2,000 construction)",
     "$118,041 avg (stated; vs county $102,817)", "N/A — stated", "$1.4B",
     "JDIG up to $19.3M / 12 yr; $6.4M IDF-Utility; +city/county", "Construction 2026; complete ~end-2028; hiring through ~2031",
     "2026-04-22", "https://news.abbvie.com/2026-04-22-AbbVie-Selects-North-Carolina-for-New-1-4-Billion-Manufacturing-Campus",
     "High", "Adv. mfg / R&D / lab", "Suburban-urban",
     "High-end MF + move-up for-sale around Durham/RTP; spillover Wake/Orange. Construction-phase rental near-term.",
     "No new in-window milestone found this cycle; project remains on track."],

    ["2026-07-06", "Updated (mixed)", 5, 81, "Greensboro / PTI Airport", "Guilford", "NC",
     "JetZero — aerospace plant (funding secured; hiring delayed)", "Aerospace / advanced manufacturing",
     "State infrastructure funding + hiring-timeline revision", 14500, "Projected",
     "$89,340 avg (stated, 2025 selection)", "N/A — stated (2025)", "$4.7B",
     "JDIG up to ~$1.02B / 37 yr; Guilford grant ~$75.9M / 20 yr; NEW: $133.9M state budget infrastructure allocation",
     "Hiring deadline pushed from 12/31/2026 to 12/31/2027; ramp now 2028-2029",
     "2026-06-30", "https://www.carolinajournal.com/state-budget-allocates-133-9m-for-delayed-jetzero-project/",
     "High (facts); Low (timeline)", "Aerospace mfg + eng.", "Suburban",
     "Same long-run scale, but push absorption modeling out ~12 months to 2028-2029.",
     "Mixed signal: state funding is a positive, but company's own hiring deadline slipped a full year due to the same budget delay."],

    ["2026-07-06", "Repeated", 6, 80, "Holly Springs / SW Wake", "Wake", "NC",
     "Genentech — biomanufacturing expansion (investment doubled to ~$2B)", "Biopharma / life sciences mfg",
     "Expansion of committed project", 500, "Projected (site total; +100 in-window, +1,500 construction)",
     "~$119,833 avg (stated, original 2025 tranche; ~$120k, 56% above Wake avg)", "N/A — stated", "~$2.0B (doubled from ~$700M)",
     "JDIG up to $9.85M / 12 yr (original 420-job award; no new incentive on the +100)", "Operational by 2029",
     "2026-01-20", "https://www.gene.com/media/press-releases/15096/2026-01-20/genentech-more-than-doubles-investment-i",
     "High", "Adv. biomanufacturing", "Suburban",
     "Strong higher-income for-sale + Class-A rental in fast-growing SW Wake (Holly Springs/Apex/Fuquay); ~$120k wage well above county median.",
     "No new in-window milestone found this cycle beyond January investment-doubling story."],

    ["2026-07-06", "Gaining Momentum", 7, 79, "Charlotte", "Mecklenburg", "NC",
     "Capital Group — East Coast operations hub (lease signed)", "Investment mgmt / finance + tech",
     "New facility / operations hub", 600, "Projected",
     "Not disclosed", "Estimated/Inferred ~$190k (from ~$116M payroll / 600 jobs; roles all $100k+ in CLT)", "$60M",
     "JDIG up to $17.2M / 12 yr; $5.7M IDF-Utility", "Leased 196,940 sq ft at One Independence Center, Uptown",
     "2026-05-28", "https://crenews.com/2026/05/28/capital-group-leases-200000-sf-at-charlottes-one-independence-center-office/",
     "High (jobs); Med (wage)", "Office — finance/tech", "Urban-suburban",
     "High-skill eng/data workforce -> walkable South End/Uptown & inner-suburban Class-A and for-sale TH.",
     "Large Uptown lease confirms the operations hub is real estate now, not just an announcement."],

    ["2026-07-06", "New (newly in-window)", 8, 77, "Durham / Morrisville", "Durham & Wake", "NC",
     "Novartis — API manufacturing building (investment increase)", "Pharmaceutical / biologics & small-molecule mfg",
     "Investment increase on existing 700-job project", 700, "Confirmed (part of pre-window 2025-11-19 package)",
     "$111,161 avg (stated; exceeds Durham avg $97,531 and Wake avg $76,643)", "N/A — stated", "~$991M (incl. new $220M)",
     "Part of original $771M package incentives", "New 56,200 sq ft API building announced 2026-04-30; facility opening 2027-2028",
     "2026-04-30", "https://www.wral.com/news/local/novartis-pharmaceuticals-expands-nc-presence-morrisville-plant-april-2026/",
     "High", "Lab/manufacturing/API production", "Suburban",
     "Reinforces sustained high-wage biomanufacturing hiring in the Wake/Durham corridor; supports continued upper-income housing absorption around Morrisville/RTP.",
     "Excluded at baseline as pre-window (orig. 2025-11-19); this in-window investment-increase milestone brings it back into scope per report rules."],
]

# ---- Week 3 (2026-07-13) ----
WEEK3_RUNNING = [
    [REPORT_DATE, "Repeated", 1, 93, "Charlotte / Uptown", "Mecklenburg", "NC",
     "SMBC Group — 2nd US HQ / bank operations", "Banking / financial services",
     "Lease signed / active hiring milestone (unchanged)", 2000, "Projected (6 yr)",
     "$165,316 avg (stated; vs county $90,706)", "N/A — stated", "$50.5M",
     "JDIG up to $70.0M / 12 yr; $23.3M IDF-Utility Account", "Leased former One Wells Fargo Center; hiring 70+ roles; first workers fall 2026",
     "2026-05-04", "https://www.axios.com/local/charlotte/2026/05/04/smbc-headquarters-wells-fargo-center-uptown",
     "High", "Office / HQ banking", "Urban",
     "Premier high-income signal; Class-A urban MF + high-end for-sale TH in Uptown/South End/Dilworth.",
     "No new in-window milestone found this cycle; project remains on track for fall 2026 first arrivals."],

    [REPORT_DATE, "Updated (mixed)", 2, 87, "Rock Hill / York Co.", "York", "SC",
     "Octapharma — \"Project Palmetto Rock\" HQ + manufacturing campus", "Biopharmaceutical (plasma fractionation)",
     "Final (3rd reading) county approval; city-level consent still pending", 1500, "Projected (1,200 new + 300 relocated)",
     "$141,502 avg HQ / $102,752 avg mfg (both stated, county-presented figures)", "N/A — stated", "$1.5B",
     "Fee-in-lieu-of-tax, assessment ratio cut to 4%; ~85% of Rock Hill's tax share redirected mostly to Rock Hill Schools",
     "York County Council 3rd reading approved 2026-07-08 (4-3, narrower than 2nd reading's 5-2); Rock Hill City Council has not yet voted on the revised terms as of 2026-07-13; Planning Commission recommended rezoning 2026-07-08",
     "2026-07-08", "https://www.wbtv.com/2026/07/10/york-county-council-approves-project-palmetto-rock-deal-rock-hill-city-council-vote-deal/",
     "High (facts); Med (deal certainty)", "Office/HQ + biopharma mfg", "Suburban",
     "Dual-tier: $141k HQ roles -> luxury for-sale/Class-A in Rock Hill/Fort Mill; $102k mfg roles -> premium workforce-plus rental/TH.",
     "County-level approval advanced but on a narrower vote; Rock Hill City Council's outstanding vote on the amended tax-share terms is the key unresolved risk heading into next cycle."],

    [REPORT_DATE, "Gaining Momentum", 3, 86, "Blythewood / N. Columbia", "Richland", "SC",
     "Scout Motors — EV assembly plant (active hiring ramp)", "Automotive / EV manufacturing",
     "Hiring event announced; companywide headcount update", 4000, "Projected (~1,400 hired companywide; 600+ in SC)",
     "$30-$37.50/hr line (stated; senior maintenance techs to $37.50/hr); salaried higher", "Estimated/Inferred mix supports above-median HH income",
     "$2.0B (+$25M training center)", "State/local package (2023); $25M readySC", "Pre-hire/interview events for 70+ maintenance-tech roles begin 2026-07-29; customer vehicle deliveries now more likely 2028 vs. earlier end-2026 framing",
     "2026-07-01", "https://www.postandcourier.com/columbia/business/scout-motors-blythewood-factory-construction-sc/article_66a10144-3603-47f1-9b23-2f8415835b8f.html",
     "High", "Auto mfg + salaried eng.", "Suburban",
     "Bifurcated: salaried/professional -> higher-income for-sale & Class-A; hourly base -> workforce/missing-middle rental in N. Richland/Fairfield.",
     "Active hiring push continues (new interview events, ~1,400 hired companywide) even as customer-delivery timeline slips toward 2028 — near-term housing-demand driver (hiring) remains strong."],

    [REPORT_DATE, "Repeated", 4, 84, "Durham / RTP", "Durham", "NC",
     "AbbVie — biopharma manufacturing campus", "Biopharma / life sciences mfg",
     "New facility (greenfield, 185 ac)", 734, "Projected (+2,000 construction)",
     "$118,041 avg (stated; vs county $102,817)", "N/A — stated", "$1.4B",
     "JDIG up to $19.3M / 12 yr; $6.4M IDF-Utility; +city/county", "Construction 2026; complete ~end-2028; hiring through ~2031",
     "2026-04-22", "https://news.abbvie.com/2026-04-22-AbbVie-Selects-North-Carolina-for-New-1-4-Billion-Manufacturing-Campus",
     "High", "Adv. mfg / R&D / lab", "Suburban-urban",
     "High-end MF + move-up for-sale around Durham/RTP; spillover Wake/Orange. Construction-phase rental near-term.",
     "No new in-window milestone found this cycle."],

    [REPORT_DATE, "Updated", 5, 81, "Charlotte", "Mecklenburg", "NC",
     "Capital Group — East Coast operations hub", "Investment mgmt / finance + tech",
     "Wage figure confirmed (stated, not inferred)", 600, "Projected",
     "$194,141 avg (stated; NC Commerce JDIG figure)", "N/A — now stated (was Estimated/Inferred ~$190k)", "$60M",
     "JDIG up to $17.2M / 12 yr; $5.7M IDF-Utility", "12-yr JDIG term; lease term began 2026-05-22",
     "2026-05-26", "https://crenews.com/2026/05/28/capital-group-leases-200000-sf-at-charlottes-one-independence-center-office/",
     "High", "Office — finance/tech", "Urban-suburban",
     "High-skill eng/data workforce -> walkable South End/Uptown & inner-suburban Class-A and for-sale TH.",
     "Wage figure upgraded from Estimated/Inferred (~$190k, payroll math) to a stated $194,141 avg (NC Commerce JDIG record) — removes prior wage-confidence discount."],

    [REPORT_DATE, "Repeated", 6, 80, "Holly Springs / SW Wake", "Wake", "NC",
     "Genentech — biomanufacturing expansion", "Adv. biomanufacturing",
     "Expansion of committed project", 500, "Projected (site total)",
     "~$119,833 avg (stated, 2025 tranche)", "N/A — stated", "~$2.0B",
     "JDIG up to $9.85M / 12 yr (original 420-job award)", "Operational by 2029",
     "2026-01-20", "https://www.gene.com/media/press-releases/15096/2026-01-20/genentech-more-than-doubles-investment-i",
     "High", "Adv. biomanufacturing", "Suburban",
     "Strong higher-income for-sale + Class-A rental in fast-growing SW Wake.",
     "No new in-window milestone found this cycle."],

    [REPORT_DATE, "Updated (mixed)", 7, 79, "Greensboro / PTI Airport", "Guilford", "NC",
     "JetZero — aerospace plant (groundbreaking held; hiring pushed further)", "Aerospace / advanced manufacturing",
     "Groundbreaking ceremony + JDIG base-period amendment", 14500, "Projected",
     "$89,340 avg (stated, 2025)", "N/A — stated (2025)", "$4.7B",
     "Transformative JDIG up to $1.018B / 37 yr; Guilford grant $75.9M / 20 yr; base period shifted to 2028-2037",
     "Groundbreaking held 2026-06-15; zero jobs now projected for 2027; hiring ramp begins 2028-2029 (3,020 jobs by end 2029); full 14,500+ target now due by 2037 (pushed from 2036)",
     "2026-06-15", "https://governor.nc.gov/news/press-releases/2026/06/15/governor-stein-celebrates-jetzero-groundbreaking-launch-greensboro-airplane-makers-14500-job-project",
     "High (facts); Low (timeline)", "Aerospace mfg + eng.", "Suburban",
     "Same long-run scale, but full-target timeline has slipped a further year (2036->2037); construction is real and tangible even as hiring pushes out.",
     "Groundbreaking actually occurred — first physical construction milestone — but the state's own JDIG amendment confirms the hiring ramp is now pushed out even further than last cycle's estimate."],

    [REPORT_DATE, "Repeated", 8, 77, "Durham / Morrisville", "Durham & Wake", "NC",
     "Novartis — API mfg building (investment increase)", "Adv. mfg / API production",
     "New facility (unchanged)", 700, "Projected (part of pre-window deal)",
     "$111,161 avg (stated)", "N/A — stated", "~$991M (incl. $220M API building)",
     "Part of original $771M package incentives", "Facility opening targeted 2027-2028",
     "2026-04-30", "https://www.wral.com/news/local/novartis-pharmaceuticals-expands-nc-presence-morrisville-plant-april-2026/",
     "High", "Adv. mfg / API production", "Suburban",
     "Reinforces sustained high-wage biomanufacturing hiring in Wake/Durham corridor.",
     "No new in-window milestone found this cycle."],
]

RUNNING = WEEK1_RUNNING + WEEK2_RUNNING + WEEK3_RUNNING

# ============================================================================
# Markets to Watch
# ============================================================================
WATCH_HEADERS = [
    "report_date", "market_submarket", "county", "state", "company_project", "job_count",
    "capital_investment", "salary_note", "why_watch", "source_date", "source_url", "confidence",
]

# ---- Week 1 (2026-06-24) ----
WEEK1_WATCH = [
    ["2026-06-24", "Cherokee Co. / Blacksburg", "Cherokee", "SC", "USA Rare Earth — rare-earth magnet plant", "~490", "$1.2B",
     "Not disclosed; 'high-skill, high-wage' (Inferred $90k-$130k+)", "Just under 500 jobs; high-wage + huge capex — strongest near-qualifier", "2026-06-02",
     "https://www.sccommerce.com/news/usa-rare-earth-inc-selects-cherokee-county-first-south-carolina-operation", "High"],
    ["2026-06-24", "Orangeburg / I-26", "Orangeburg", "SC", "Ferrara Candy — new mfg + corporate site", "1,000 (10 yr)", "$675M",
     "Not disclosed; confectionery = workforce-tier", "500+ but fails high-income intent; workforce-housing relevance", "2026-04-22",
     "https://governor.sc.gov/news/2026-04/ferrara-candy-company-selects-orangeburg-county-first-south-carolina-operation", "High (facts)"],
    ["2026-06-24", "Laurens Co.", "Laurens", "SC", "Suniva — solar-cell plant", "564", "$350M",
     "Stated $23-$53/hr = workforce-tier", "500+ but wage below $100k median; workforce housing", "2026-04-14",
     "https://governor.sc.gov/news/2026-04/suniva-inc-selects-laurens-county-first-south-carolina-manufacturing-facility", "High"],
    ["2026-06-24", "Columbia / BullStreet", "Richland", "SC", "AMAROK — new HQ (perimeter security)", "296", "$69M",
     "Not disclosed; HQ/professional (portion likely $100k+)", "Sub-500 but best $100k-friendly Midlands signal", "2026-03",
     "https://governor.sc.gov/news/2026-03/amarok-expands-richland-county-operations-new-headquarters", "Med"],
    ["2026-06-24", "Greensboro", "Guilford", "NC", "Lumentum — AI/data-center optics", "~400", "Hundreds of millions",
     "Not disclosed", "Sub-500; AI-supply-chain relevance", "2026-04-14",
     "https://www.areadevelopment.com/newsitems/4-14-2026/lumentum-greensboro-north-carolina.shtml", "Med"],
    ["2026-06-24", "Hendersonville", "Henderson", "NC", "BorgWarner — vertical-integration expansion", "378", "$100M",
     "$67,047 avg (stated; below $100k)", "Sub-500; solid WNC manufacturing signal", "2026-05-26",
     "https://www.commerce.nc.gov/news/press-releases/2026/05/26/governor-stein-announces-100-million-expansion-borgwarner-hendersonville", "High"],
    ["2026-06-24", "Asheville / Arden", "Buncombe", "NC", "Eaton — Low Voltage Assembly expansion", "300", "Not disclosed",
     "'high-wage' but figure Not disclosed (Inferred mid-$60k-$80k)", "Sub-500; post-Helene WNC cluster", "2026-04-07",
     "https://www.ashevillechamber.org/news-events/press-releases/eaton-invests-in-workforce-growth-in-buncombe-county/", "High (jobs); Low (wage)"],
    ["2026-06-24", "Charlotte / Raleigh / Rural Hall", "Multi", "NC", "Siemens Energy — $421M NC expansion", "500 (statewide)", "$421M",
     "Not disclosed (Inferred ~$87k from prior tranche; sub-$100k)", "In-window (2026-02-04) but 500 jobs split across 3 metros, no single 500+ site; wage sub-$100k", "2026-02-04",
     "https://businessnc.com/siemens-energys-421-million-n-c-expansion-adding-500-jobs/", "High (totals); Low (wage/split)"],
    ["2026-06-24", "Knightdale / Wendell", "Wake", "NC", "Siemens AG — power devices for AI/data centers", "350", "part of $165M",
     "Not disclosed", "In-window (2026-03-17) but sub-500; mfg wages likely sub-$100k", "2026-03-17",
     "https://www.wral.com/business/siemens-350-jobs-nc-sc-165m-investment-ai-data-centers-raleigh-wendell-march-2026/", "High"],
    ["2026-06-24", "Kernersville", "Forsyth", "NC", "John Deere — excavator plant (relocating from Japan)", "150+", "$70M",
     "Not disclosed (Inferred skilled-mfg)", "In-window (Jan 2026) but sub-500; marquee Triad tenant", "2026-01",
     "https://myfox8.com/news/north-carolina/piedmont-triad/john-deere-factory-bringing-150-jobs-to-piedmont-triad/", "Med"],
    ["2026-06-24", "Spartanburg Co.", "Spartanburg", "SC", "TigerDC 'Project Spero' — data center", "~50 FTE (Phase I)", "$3B",
     "Not disclosed", "In-window (2026-01-27); enormous capex but very few permanent jobs", "2026-01-27",
     "https://www.upstatescalliance.com/data-resources/media-center/", "Med (capex)"],
    ["2026-06-24", "York Co.", "York", "SC", "QTS 'Project Cobra' — data-center campus", "~200 FTE (+~1,000 constr.)", "$1B+ (up to $8B)",
     "~$80k median (stated)", "Charlotte-metro SC target county; massive capex, few permanent jobs", "2026-01 (buildout)",
     "https://www.postandcourier.com/york-county/news/qts-data-center-york-county-community-meeting/article_fb2c12f3-d1a8-41eb-bd58-d89d8dfb92d5.html", "High (facts)"],
    ["2026-06-24", "Indian Land", "Lancaster", "SC", "Snider Fleet Solutions — corporate office relo", "167", "$6.9M",
     "Not disclosed; corporate/HQ roles", "Sub-500 but relo into high-growth Charlotte-metro SC submarket", "2026",
     "https://www.qcnews.com/news/u-s/lancaster-county/major-manufacturing-company-moving-to-indian-land/", "Med"],
    ["2026-06-24", "Berkeley / Dorchester", "Berkeley & Dorchester", "SC", "Google — data-center expansion", "~160 apprentices (FTE n/d)", "$9B",
     "Not disclosed", "Huge capex but announced Oct 2025 (pre-window) + jobs undisclosed", "2025-10-13",
     "https://blog.google/company-news/inside-google/company-announcements/google-american-innovation-south-carolina/", "High (capex)"],
    ["2026-06-24", "Liberty", "Pickens", "SC", "FN America — 2nd production facility", "~176", "$33M",
     "Not disclosed", "Sub-500; Upstate target county", "2026",
     "https://www.sccommerce.com/news/fn-america-llc-expanding-south-carolina-footprint-pickens-county-second-production-facility", "High"],
    ["2026-06-24", "Spartanburg", "Spartanburg", "SC", "Siemens Smart Infrastructure", "~150", "$165M",
     "Not disclosed", "Sub-500; Upstate target county", "2026-03-18",
     "https://www.foxcarolina.com/2026/03/18/tech-manufacturer-building-expanding-upstate-facilities-creating-150-new-jobs/", "Med-High"],
    ["2026-06-24", "Hamlet", "Richmond", "NC", "AWS/Amazon — AI/cloud campus", "500", "$10B",
     "Not disclosed", "NC + 500 jobs + huge capex, but rural & far from footprint; VERIFY date/wage", "verify",
     "https://www.aboutamazon.com/", "Low (needs verification)"],
    ["2026-06-24", "Wake Co. (aggregate)", "Wake", "NC", "County jobs pipeline (52 projects)", "~11,000 (pipeline)", "$11B",
     "Mixed (office-skewed)", "Aggregate pipeline, not a single project; downtown Raleigh office signal", "2026-02",
     "https://nchospitalityalliance.com/wake-county-pursues-11-billion-jobs-pipeline/", "Med"],
]

# ---- Week 2 (2026-07-06) — carried-forward items updated + new items ----
WEEK2_WATCH = [
    ["2026-07-06", "Cherokee Co. / Blacksburg", "Cherokee", "SC", "USA Rare Earth — rare-earth magnet plant", "~490", "$1.2B",
     "$24.50-$63/hr (newly stated)", "Still just under 500 jobs; wage newly disclosed this cycle", "2026-06-02",
     "https://scdailygazette.com/2026/06/02/", "High"],
    ["2026-07-06", "Orangeburg / I-26", "Orangeburg", "SC", "Ferrara Candy — groundbreaking held", "1,000 (10 yr)", "$675M",
     "Not disclosed; confectionery = workforce-tier", "Groundbreaking held 2026-05-20; still fails high-income intent", "2026-05-20",
     "https://www.wrdw.com/2026/05/20/ferrara-candy-company-breaks-ground-orangeburg/", "High"],
    ["2026-07-06", "Laurens Co.", "Laurens", "SC", "Suniva — solar-cell plant", "564", "$350M",
     "Stated $23-$53/hr = workforce-tier", "No update found this cycle; unchanged workforce-tier read", "2026-04-14",
     "https://governor.sc.gov/news/2026-04/suniva-inc-selects-laurens-county-first-south-carolina-manufacturing-facility", "High"],
    ["2026-07-06", "Columbia / BullStreet", "Richland", "SC", "AMAROK — new HQ (perimeter security)", "296", "$69M",
     "Not disclosed; HQ/professional (portion likely $100k+)", "Richland County Council approved incentive package 2026-03-03", "2026-03-03",
     "https://www.postandcourier.com/columbia/business/columbia-sc-bullstreet-amarok-headquarters/article_73ea5904-fe8c-4048-b0a3-10417d1af35e.html", "High"],
    ["2026-07-06", "Greensboro", "Guilford", "NC", "Lumentum — AI/data-center optics", "~400", "Hundreds of millions",
     "Not disclosed", "No update found this cycle", "2026-04-14",
     "https://www.areadevelopment.com/newsitems/4-14-2026/lumentum-greensboro-north-carolina.shtml", "Med"],
    ["2026-07-06", "Hendersonville", "Henderson", "NC", "BorgWarner — vertical-integration expansion", "378", "$100M",
     "$67,047 avg (stated; below $100k)", "A 2026-05-27 follow-up mention was found but not corroborated with detail this cycle — recheck next report", "2026-05-27",
     "https://www.commerce.nc.gov/news/press-releases/2026/05/26/governor-stein-announces-100-million-expansion-borgwarner-hendersonville", "Med"],
    ["2026-07-06", "Asheville / Arden", "Buncombe", "NC", "Eaton — Low Voltage Assembly expansion", "300", "Not disclosed",
     "'high-wage' but figure Not disclosed (Inferred mid-$60k-$80k)", "Confirmed as the 7th Helene-recovery manufacturer expansion (725+ cumulative jobs, $388M+ across 7 deals)", "2026-04-07",
     "https://www.ashevillechamber.org/news-events/press-releases/eaton-invests-in-workforce-growth-in-buncombe-county/", "High (jobs); Low (wage)"],
    ["2026-07-06", "Charlotte / Raleigh / Rural Hall", "Multi", "NC", "Siemens Energy — $421M NC expansion", "500 (statewide)", "$421M",
     "Not disclosed (Inferred ~$87k from prior tranche; sub-$100k)", "No update found this cycle", "2026-02-04",
     "https://businessnc.com/siemens-energys-421-million-n-c-expansion-adding-500-jobs/", "High (totals); Low (wage/split)"],
    ["2026-07-06", "Raleigh / Wendell / Knightdale", "Wake", "NC", "Siemens AG — power devices for AI/data centers", "350", "part of $165M",
     "Not disclosed", "Site detail confirmed: 100 Raleigh by YE2026, 50 new Wendell site, 200+ Wendell HQ by 2028", "2026-03-17",
     "https://www.wral.com/business/siemens-350-jobs-nc-sc-165m-investment-ai-data-centers-raleigh-wendell-march-2026/", "High"],
    ["2026-07-06", "Kernersville", "Forsyth", "NC", "John Deere — excavator plant (relocating from Japan)", "150+", "$70M",
     "Not disclosed (Inferred skilled-mfg)", "Re-confirmed Jan 2026 w/ White House mention; construction later 2026, production 2030", "2026-01",
     "https://myfox8.com/news/north-carolina/piedmont-triad/john-deere-factory-bringing-150-jobs-to-piedmont-triad/", "High"],
    ["2026-07-06", "Spartanburg Co.", "Spartanburg", "SC", "TigerDC 'Project Spero' — data center", "~50 FTE (Phase I)", "$3B",
     "~$100,000 avg (newly stated)", "Wage newly disclosed this cycle (~2x county avg)", "2026-01-27",
     "https://www.upstatescalliance.com/data-resources/media-center/", "Med-High (capex/wage)"],
    ["2026-07-06", "York Co.", "York", "SC", "QTS 'Project Cobra' — data-center campus", "~200 FTE (+~1,000 constr.)", "up to $8B",
     "~$80k median (stated)", "Investment grown to ~$8B; June 2026 county moratorium discussion tied partly to this project", "2026-06",
     "https://www.postandcourier.com/york-county/news/qts-data-center-york-clover-meeting/article_d3095fdb-8572-4510-9974-681459f18cd9.html", "Med (capex figure varies by source)"],
    ["2026-07-06", "Indian Land", "Lancaster", "SC", "Snider Fleet Solutions — corporate office relo", "167", "$6.9M",
     "Not disclosed; corporate/HQ roles", "No update found this cycle", "2026",
     "https://www.qcnews.com/news/u-s/lancaster-county/major-manufacturing-company-moving-to-indian-land/", "Med"],
    ["2026-07-06", "Berkeley / Dorchester", "Berkeley & Dorchester", "SC", "Google — data-center expansion", "~160 apprentices (FTE n/d)", "$9B",
     "Not disclosed", "No new in-window milestone found this cycle", "2025-10-13",
     "https://blog.google/company-news/inside-google/company-announcements/google-american-innovation-south-carolina/", "High (capex)"],
    ["2026-07-06", "Liberty", "Pickens", "SC", "FN America — 2nd production facility", "~176", "$33M",
     "Not disclosed", "No 2026 milestone found; appears to have gone quiet — Losing Relevance", "2026",
     "https://www.sccommerce.com/news/fn-america-llc-expanding-south-carolina-footprint-pickens-county-second-production-facility", "Low"],
    ["2026-07-06", "Spartanburg", "Spartanburg", "SC", "Siemens Smart Infrastructure", "~150", "$165M",
     "Not disclosed", "No update found this cycle", "2026-03-18",
     "https://www.foxcarolina.com/2026/03/18/tech-manufacturer-building-expanding-upstate-facilities-creating-150-new-jobs/", "Med-High"],
    ["2026-07-06", "Hamlet", "Richmond", "NC", "AWS/Amazon — AI/cloud campus", "500", "$10B",
     "Not disclosed", "Still unverified; needs direct confirmation", "verify",
     "https://www.aboutamazon.com/", "Low (needs verification)"],
    ["2026-07-06", "Wake Co. (aggregate)", "Wake", "NC", "County jobs pipeline (52 projects)", "~11,000 (pipeline)", "$11B",
     "Mixed (office-skewed)", "A second, possibly overlapping aggregate figure (~30 projects/8,000 jobs/$5.6B) surfaced this cycle — flagged for reconciliation", "2026-02",
     "https://nchospitalityalliance.com/wake-county-pursues-11-billion-jobs-pipeline/", "Med"],
    ["2026-07-06", "Charlotte / SouthPark", "Mecklenburg", "NC", "JPMorgan Chase — new office (1,000-employee bldg)", "400 new by 2028", "Not disclosed",
     "~$105k Estimated/Inferred (PayScale avg)", "Sub-500 new jobs; wage inferred", "2026-04-21",
     "https://www.axios.com/local/charlotte/2026/04/21/jpmorganchase-charlotte-hiring-workforce-banking-competition", "Med-High"],
    ["2026-07-06", "Charlotte (CLT airport)", "Mecklenburg", "NC", "Averitt — logistics campus", "211", "$200M",
     "$81,769 avg (stated)", "Sub-500; wage below $100k", "2026-04-29",
     "https://news.mecknc.gov/averitt-announces-major-new-commitment-mecklenburg-county", "High"],
    ["2026-07-06", "Rowan County", "Rowan", "NC", "\"Project Rack\" (company undisclosed)", "258", "$41M",
     "'Above-average starting' Estimated/Inferred sub-$100k", "Sub-500; wage tier unclear; company undisclosed", "2026-06-06",
     "https://www.salisburypost.com/2026/06/06/rowan-county-approves-project-rack-incentives-for-258-job-distribution-site/", "Med"],
    ["2026-07-06", "Rowan / Kannapolis", "Rowan", "NC", "Google (via DHL) — 730k sq ft warehouse lease", "Not disclosed", "Not disclosed",
     "$55k-$150k range (DHL postings)", "Jobs undisclosed; capex-heavy lease", "2026-04-29",
     "https://www.wbtv.com/2026/04/29/google-leases-massive-facility-near-rowan-cabarrus-county-line/", "Med"],
    ["2026-07-06", "Raleigh / North Hills", "Wake", "NC", "Ralliant Corp. — global HQ launch", "180", "$2.1M",
     "$170k-$189k avg (stated)", "Sub-500 but very high wage; HQ opened March 2026, active hiring ramp", "2026-03",
     "https://www.wral.com/business/technology/ralliant-launches-raleigh-global-hq-north-hills-march-2026/", "High"],
    ["2026-07-06", "Siler City", "Chatham", "NC", "Wolfspeed — silicon-carbide fab", "1,800 (proj.)", "$5.0B",
     "$77,753 avg (stated)", "500+ jobs but wage sub-$100k; full occupancy targeted March 2026, production June 2026, 200+ hired", "2026-03",
     "https://www.chathamedc.org/news/wolfspeed-siler-city/", "Med"],
    ["2026-07-06", "Benson", "Johnston", "NC", "Vulcan Elements — magnet factory", "1,000 (proj.)", "$918.4M",
     "$81,932 avg (stated)", "500+ jobs but wage sub-$100k; JCC training cohorts begin summer 2026", "2026-summer (training start)",
     "https://www.wunc.org/economy/2025-11-18/vulcan-elements-worlds-largest-magnet-factory-outside-china-benson", "Med"],
    ["2026-07-06", "Wake County (statewide)", "Wake", "NC", "WakeMed / Atrium Health — proposed merger", "3,300 (statewide, 5 yr)", "$2.0B (Wake Co. portion)",
     "Not disclosed", "Huge but wage/geography unclear; still under regulatory/public review", "2026-05-01",
     "https://abc11.com/post/nc-wakemed-atrium-merge-creating-2b-investment-wake-county-3300-health-care-jobs/19016838/", "Med"],
    ["2026-07-06", "Greensboro / PTI Airport", "Guilford", "NC", "Boom Supersonic — Overture assembly (AT RISK)", "<=500 currently; must hit 500 by 12/31/26", "~$500M (site, existing)",
     "Not disclosed", "RISK FLAG: hangar reportedly largely idle; contractual hiring cliff this year or lease may terminate", "2026-03",
     "https://hoodline.com/2026/03/greensboro-s-50-million-supersonic-ghost-hangar-puts-state-on-the-clock/", "Med"],
    ["2026-07-06", "Multi-site NC incl. Wilmington", "New Hanover +", "NC", "Amazon-Corning — fiber-optics partnership", "1,000 (multi-site, unallocated)", "Multi-billion (undisclosed NC-specific)",
     "Not disclosed", "Wilmington-specific job count not disclosed", "2026-06-08",
     "https://www.wilmingtonbiz.com/technology/2026/06/08/amazon_corning_strike_multibillion_dollar_deal_to_add_1000_nc_jobs/27560", "Med"],
    ["2026-07-06", "Arden", "Buncombe", "NC", "Pratt & Whitney (RTX) — casting foundry expansion", "325", "$285M",
     "$62,413 avg (stated)", "Sub-$100k wage; equipment arrival end 2026, first parts mid-2027", "2025-01 (orig.); 2026 milestone",
     "https://businessnc.com/pratt-whitney-announces-285-million-expansion-325-more-jobs-for-buncombe-county/", "Med"],
    ["2026-07-06", "Asheboro", "Randolph", "NC", "Environmental Air Systems", "300", "$20M",
     "$55,133 avg (stated)", "Sub-$100k wage", "2025-11-25 (orig.)",
     "https://businessnc.com/randolph-county-wins-20-million-hvac-project-300-jobs/", "Med"],
    ["2026-07-06", "Wilmington", "New Hanover", "NC", "GE Hitachi / Global Nuclear Fuel — GNF4 fuel line", "Not disclosed", "Not itemized",
     "Not disclosed", "Jobs undisclosed for this milestone (production start early 2026)", "2026-01 (est.)",
     "https://www.neimagazine.com/news/ge-hitachi-allocates-50m-to-enhance-safety-at-north-carolina-factory/", "Low-Med"],
    ["2026-07-06", "Woodruff / Spartanburg Co.", "Spartanburg", "SC", "AIRSYS Cooling Technologies — global HQ", "215", "$40-60M",
     "Not disclosed (Estimated/Mixed)", "Sub-500; wage not disclosed", "2026-05-13",
     "https://www.airsysnorthamerica.com/", "High"],
    ["2026-07-06", "Gray Court", "Laurens", "SC", "Aptiv Services US — Connexial Center", "277", "$120.8M",
     "Not disclosed", "Sub-500; wage not disclosed", "2026-04-20",
     "https://www.growlaurenscounty.com/", "Med"],
    ["2026-07-06", "Easley (Anderson/Pickens border)", "Anderson", "SC", "Signature Foods USA", "202", "$11.5M",
     "Estimated/Inferred sub-$50k (food mfg.)", "Sub-100k wage; county border ambiguous (Anderson per state, Pickens per local coverage)", "2026-04",
     "https://www.wltx.com/", "Med"],
    ["2026-07-06", "Greenville", "Greenville", "SC", "GNQ Insilico — AI/quantum techbio HQ", "Not disclosed", "$500M (valuation, not capex)",
     "Not disclosed", "Jobs entirely undisclosed; company 'still developing projections'", "2026-05-23",
     "https://www.scbio.org/", "Med"],
    ["2026-07-06", "North Charleston", "Charleston", "SC", "Boeing 787 — production-rate ramp", "1,000 (unchanged)", "$1.0B (existing)",
     "Estimated/Inferred mixed ($54-64k assemblers; eng. likely higher)", "Rate ramp 7->10/month during 2026 confirms hiring plan active; wage mixed", "2026-01-27",
     "https://aviationweek.com/air-transport/aircraft-propulsion/boeing-formally-launches-787-facilities-expansion", "Med"],
    ["2026-07-06", "North Charleston", "Charleston", "SC", "SC Ports Authority — Leatherman Terminal Phase 2", "400 direct + ~1,200 indirect", "Not itemized",
     "Not disclosed (Estimated mixed, some 6-figure specialist roles)", "Direct count sub-500; wage unclear", "2026-04-01",
     "https://www.joc.com/article/charlestons-leatherman-terminal-eyes-second-berth-by-late-2026-6066963", "Med"],
    ["2026-07-06", "Goose Creek", "Berkeley", "SC", "HII (Newport News Shipbuilding) — 1-yr anniversary", "~250 (estimate)", "Not disclosed",
     "Not disclosed", "Sub-500; figure not re-confirmed this cycle", "2026-01-23",
     "https://www.berkeleyobserver.com/", "Low"],
]

# ---- Week 3 (2026-07-13) — carried-forward items updated + new items; TigerDC removed (withdrawn) ----
WEEK3_WATCH = [
    [REPORT_DATE, "Cherokee Co. / Blacksburg", "Cherokee", "SC", "USA Rare Earth — rare-earth magnet plant", "~490", "$1.2B",
     "$24.50-$63/hr (stated; unchanged)", "Still just under 500 jobs; no threshold-crossing update found this cycle (high-priority check)", "2026-06-02",
     "https://www.postandcourier.com/spartanburg/business/usa-rare-earth-blacksburg-sc-facility/article_2681ca76-7e5f-4a73-8a80-13231e38917f.html", "High"],
    [REPORT_DATE, "Orangeburg / I-26", "Orangeburg", "SC", "Ferrara Candy — groundbreaking held", "1,000 (10 yr)", "$675M",
     "Not disclosed; confectionery = workforce-tier", "No new milestone since May 2026 groundbreaking; first lines targeted Q1 2029", "2026-05-20",
     "https://www.wrdw.com/2026/05/20/ferrara-candy-company-breaks-ground-orangeburg/", "High"],
    [REPORT_DATE, "Laurens Co.", "Laurens", "SC", "Suniva — solar-cell plant", "564", "$350M",
     "Stated $23-$53/hr = workforce-tier", "No update found this cycle; operations targeted spring/summer 2027", "2026-04-14",
     "https://www.postandcourier.com/greenville/business/suniva-solar-cell-plant-laurens-sc/article_986c2fc2-2335-48cb-bd3c-92e89604c49c.html", "High"],
    [REPORT_DATE, "Columbia / BullStreet", "Richland", "SC", "AMAROK — new HQ (perimeter security)", "296", "$69M",
     "Not disclosed; HQ/professional (portion likely $100k+)", "No update found this cycle", "2026-03-03",
     "https://www.postandcourier.com/columbia/business/columbia-sc-bullstreet-amarok-headquarters/article_73ea5904-fe8c-4048-b0a3-10417d1af35e.html", "High"],
    [REPORT_DATE, "Greensboro", "Guilford", "NC", "Lumentum — AI/data-center optics", "~400", "Hundreds of millions",
     "Not disclosed", "Ribbon-cutting/opening ceremony held late April 2026 for the 240,000 sq ft facility; production ramp still targeted mid-2028", "2026-04",
     "https://myfox8.com/news/north-carolina/greensboro/lumentum-holdings-opens-new-manufacturing-facility-in-greensboro/", "High"],
    [REPORT_DATE, "Hendersonville", "Henderson", "NC", "BorgWarner — vertical-integration expansion", "378", "$100M",
     "$67,047 avg (stated; below $100k)", "Now corroborated via official NC Commerce/Governor press releases dated 2026-05-26, resolving prior cycle's uncorroborated flag", "2026-05-26",
     "https://www.commerce.nc.gov/news/press-releases/2026/05/26/governor-stein-announces-100-million-expansion-borgwarner-hendersonville", "High"],
    [REPORT_DATE, "Asheville / Arden", "Buncombe", "NC", "Eaton — Low Voltage Assembly expansion", "300", "Not disclosed",
     "'high-tech, high-wage' but figure Not disclosed (Inferred mid-$60k-$80k)", "Eaton recruiting at WNC Career Expo 2026-04-16; confirmed as 7th Helene-recovery manufacturer expansion", "2026-04-08",
     "https://www.ashevillechamber.org/news-events/press-releases/eaton-invests-in-workforce-growth-in-buncombe-county/", "High (jobs); Low (wage)"],
    [REPORT_DATE, "Charlotte / Raleigh / Rural Hall", "Multi", "NC", "Siemens Energy — $421M NC expansion", "500 (statewide)", "$421M",
     "Not disclosed (Inferred ~$87k from prior tranche; sub-$100k)", "No update found this cycle; Mecklenburg-specific allocation for this tranche remains unclear (conflated with a 2024 tranche in some coverage)", "2026-02-04",
     "https://businessnc.com/siemens-energys-421-million-n-c-expansion-adding-500-jobs/", "Med (totals uncertain)"],
    [REPORT_DATE, "Raleigh / Wendell / Knightdale", "Wake", "NC", "Siemens AG — power devices for AI/data centers", "350", "part of $165M",
     "Not disclosed", "New 101,000 sq ft Wendell facility coming online July 2026; Knightdale targeting 100 jobs by end 2026; expanded Wendell campus (+200 jobs) by 2028", "2026-03-17",
     "https://www.wral.com/business/siemens-350-jobs-nc-sc-165m-investment-ai-data-centers-raleigh-wendell-march-2026/", "High"],
    [REPORT_DATE, "Kernersville", "Forsyth", "NC", "John Deere — excavator plant (relocating from Japan)", "150+", "$70M",
     "Not disclosed (Inferred skilled-mfg)", "Reaffirmed 2026-01-28 via White House-linked event; construction underway on west campus, facility slated to open within the next year", "2026-01-28",
     "https://greensboro.com/news/local/business/development/article_f8077494-ba1d-526b-a82a-9d8425f0acbe.html", "High"],
    [REPORT_DATE, "York Co.", "York", "SC", "QTS \"Project Cobra\" — data-center campus", "~200 FTE (+~1,000 constr.)", "up to $8B",
     "~$80k median (stated)", "York County's proposed nine-month data-center moratorium reached 3rd reading/public hearing 2026-07-13 (same day as this report); outcome not yet known — could affect QTS's remaining unpermitted buildings", "2026-07-13",
     "https://www.wrhi.com/2026/06/york-county-council-approves-data-center-moratorium-first-reading-advances-1-5-billion-biopharmaceutical-project-213500", "Med (regulatory outcome pending)"],
    [REPORT_DATE, "Indian Land", "Lancaster", "SC", "Snider Fleet Solutions — corporate office relo", "167", "$6.9M",
     "Not disclosed; corporate/HQ roles", "Re-examined this cycle: underlying announcement traces to 2023 with no 2026 developments found at all — likely a stale carryover; recommend removal next cycle absent a genuine update", "2023 (orig.)",
     "https://www.qcnews.com/news/u-s/lancaster-county/major-manufacturing-company-moving-to-indian-land/", "Low (stale)"],
    [REPORT_DATE, "Berkeley / Dorchester", "Berkeley & Dorchester", "SC", "Google — data-center expansion", "~160 apprentices (FTE n/d)", "$9B",
     "Not disclosed", "No new in-window milestone found this cycle; original Oct. 2025 announcement remains the latest news", "2025-10-13",
     "https://blog.google/company-news/inside-google/company-announcements/google-american-innovation-south-carolina/", "High (capex)"],
    [REPORT_DATE, "Liberty", "Pickens", "SC", "FN America — 2nd production facility", "~176", "$33M",
     "Not disclosed", "Second consecutive cycle with no momentum found; flag for possible removal next report", "2026",
     "https://www.sccommerce.com/news/fn-america-llc-expanding-south-carolina-footprint-pickens-county-second-production-facility", "Low"],
    [REPORT_DATE, "Spartanburg", "Spartanburg", "SC", "Siemens Smart Infrastructure", "~150", "$165M",
     "Not disclosed", "No update found this cycle", "2026-03-18",
     "https://www.foxcarolina.com/2026/03/18/tech-manufacturer-building-expanding-upstate-facilities-creating-150-new-jobs/", "Med-High"],
    [REPORT_DATE, "Hamlet", "Richmond", "NC", "AWS/Amazon — AI/cloud campus", "500", "$10B",
     "Not disclosed", "Still unverified; outside this cycle's five assigned research regions — needs direct confirmation next cycle", "verify",
     "https://www.aboutamazon.com/", "Low (needs verification)"],
    [REPORT_DATE, "Wake Co. (aggregate)", "Wake", "NC", "County jobs pipeline (EDGE 7, 52 projects)", "~11,000 (pipeline)", "$11B",
     "Mixed (office-skewed)", "RECONCILED this cycle: the previously-flagged second figure (~30 projects/8,000-10,800 jobs/$5.6B) is confirmed to be the closed, prior EDGE 6 campaign (concluded ~Sept 2024) — sequential, not duplicative, with the current $11B/11,000-job EDGE 7 pipeline", "2026-02",
     "https://nchospitalityalliance.com/wake-county-pursues-11-billion-jobs-pipeline/", "Med"],
    [REPORT_DATE, "Charlotte / SouthPark", "Mecklenburg", "NC", "JPMorgan Chase — new office (1,000-employee bldg)", "400 new by 2028", "Not disclosed",
     "~$105k Estimated/Inferred (job-posting range)", "No update found this cycle", "2026-04-21",
     "https://www.axios.com/local/charlotte/2026/04/21/jpmorgan-southpark-charlotte-400-jobs", "Med-High"],
    [REPORT_DATE, "Charlotte (CLT airport)", "Mecklenburg", "NC", "Averitt — logistics campus", "211", "$200M",
     "$81,769 avg (stated)", "No update found this cycle", "2026-04-29",
     "https://www.axios.com/local/charlotte/2026/04/29/averitt-charlotte-douglas-airport-campus-jobs", "High"],
    [REPORT_DATE, "Rowan County", "Rowan", "NC", "\"Project Rack\" (company undisclosed)", "258", "$41M",
     "'Above-average starting' Estimated/Inferred sub-$100k", "No update found this cycle; Q1 2027 operations start still targeted", "2026-06-06",
     "https://www.salisburypost.com/2026/06/06/rowan-county-approves-project-rack-incentives-for-258-job-distribution-site/", "Med"],
    [REPORT_DATE, "Rowan / Kannapolis", "Rowan", "NC", "Google (via DHL) — 730k sq ft warehouse lease", "Not disclosed", "Not disclosed",
     "$55k-$150k range (DHL postings)", "No update found this cycle; jobs still undisclosed", "2026-04-29",
     "https://www.salisburypost.com/2026/05/02/not-a-data-center-google-leases-rowan-county-side-kannapolis-industrial-site/", "Med"],
    [REPORT_DATE, "Raleigh / North Hills", "Wake", "NC", "Ralliant Corp. — global HQ launch", "180", "$2.1M",
     "$170k-$189k avg (stated)", "No update found this cycle", "2026-03",
     "https://www.wral.com/business/technology/ralliant-launches-raleigh-global-hq-north-hills-march-2026/", "High"],
    [REPORT_DATE, "Siler City", "Chatham", "NC", "Wolfspeed — silicon-carbide fab", "1,800 (proj.)", "$5.0B",
     "$77,753 avg (stated)", "Post-bankruptcy ramp described by the company as 'more measured/disciplined' — original full-occupancy March 2026 target appears to have slipped; SiC crystal growth/wafer shipments to Mohawk Valley (NY) underway", "2026-Q1 (10-Q)",
     "https://www.sec.gov/Archives/edgar/data/0000895419/000089541926000030/wolf-20260329.htm", "Med (financial distress clouds trajectory)"],
    [REPORT_DATE, "Benson", "Johnston", "NC", "Vulcan Elements — magnet factory", "1,000 (proj.)", "$918.4M",
     "$81,932 avg (stated)", "'Hundreds of workers already on site' per current coverage; JCC training cohorts begin summer 2026 as originally scheduled", "2026-summer (training start)",
     "https://www.wral.com/sponsored/fifth-third-bank/the-next-industrial-revolution-how-vulcan-elements-makes-nc-the-backbone-of-vital-global-market/", "Med"],
    [REPORT_DATE, "Wake County (statewide)", "Wake", "NC", "WakeMed / Atrium Health — proposed merger", "3,300 (statewide, 5 yr)", "$2.0B (Wake Co. portion)",
     "Not disclosed", "Wake Co. Commissioners' vote (delayed 90 days from 2026-05-04, targeted ~early Aug 2026) still not scheduled as of this report; opposition mounting (cost-impact study, Patients Union, NC Justice Center, NAACP)", "2026-05-04 (delay announced)",
     "https://www.wral.com/news/local/wakemed-atrium-merger-costs-opposition-july-2026/", "Med"],
    [REPORT_DATE, "Greensboro / PTI Airport", "Guilford", "NC", "Boom Supersonic — Overture assembly (AT RISK)", "<=500 currently; must hit 500 by 12/31/26 or lease may terminate", "~$50M (hangar, existing)",
     "Not disclosed", "RISK ESCALATING: hangar still largely idle as of June 2026; CEO says production starts 'in about two years'; company separately pursuing a Colorado data-center initiative, raising questions about Greensboro focus; deadline now 6 months away with no positive movement", "2026-06",
     "https://www.prismnews.com/local/guilford-nc/boom-supersonics-greensboro-factory-sits-idle-as-state", "High (risk)"],
    [REPORT_DATE, "Multi-site NC incl. Wilmington", "New Hanover +", "NC", "Amazon-Corning — fiber-optics partnership", "1,000 (multi-site, unallocated)", "Multi-billion (undisclosed NC-specific)",
     "Avg salary >$65,000 (newly disclosed; sub-$100k)", "Wage now disclosed (exceeds $65k) but still sub-$100k; per-site (incl. Wilmington-specific) job allocation still not broken out", "2026-06-08",
     "https://www.corning.com/worldwide/en/about-us/news-events/news-releases/2026/06/amazon-announces-agreement-with-corning-to-boost-us-fiber-optics-manufacturing-creating-1000-advanced-manufacturing-jobs-in-north-carolina.html", "Med"],
    [REPORT_DATE, "Arden", "Buncombe", "NC", "Pratt & Whitney (RTX) — casting foundry expansion", "325", "$285M",
     "$62,413 avg (stated)", "On track: first equipment arrival end of 2026, first parts production mid-2027, hiring ramp begins mid-2027", "2026 (equipment arrival timeline)",
     "https://www.flightglobal.com/engines/2026/04/rtx-says-gtf-groundings-are-coming-down-thanks-to-maintenance-ramp/", "Med-High"],
    [REPORT_DATE, "Asheboro", "Randolph", "NC", "Environmental Air Systems", "300", "$20M",
     "$55,133 avg (stated)", "No update found this cycle despite targeted searches", "2025-11-25 (orig.)",
     "https://businessnc.com/randolph-county-wins-20-million-hvac-project-300-jobs/", "Med"],
    [REPORT_DATE, "Wilmington", "New Hanover", "NC", "GE Hitachi / Global Nuclear Fuel — GNF4 fuel line", "Not disclosed", "Not itemized",
     "Not disclosed", "GNF4 fuel fabrication began early 2026; lead-use assemblies scheduled for deployment in commercial reactors during 2026 — genuine in-window production milestones, jobs still undisclosed", "2026 (fabrication start)",
     "https://www.world-nuclear-news.org/articles/global-nuclear-fuel-unveils-new-gnf4-product", "Med"],
    [REPORT_DATE, "Woodruff / Spartanburg Co.", "Spartanburg", "SC", "AIRSYS Cooling Technologies — global HQ", "215", "$40-60M",
     "Not disclosed (Estimated/Mixed)", "Manufacturing-operations start slipped from 2026 to early 2027; HQ campus itself already opened", "~2026 (HQ opening); early 2027 (mfg slip)",
     "https://www.postandcourier.com/spartanburg/business/airsys-data-center-water-woodruff-sc/article_86465e32-7d21-4bba-9a9b-3e6ea0e2d691.html", "Medium"],
    [REPORT_DATE, "Gray Court", "Laurens", "SC", "Aptiv Services US — Connexial Center", "277", "$120.8M",
     "Not disclosed", "No update found this cycle", "2026-04-20",
     "https://www.golaurens.com/news/laurens-county-welcomes-277-jobs-and-120-million-from-aptiv/article_a0c4713d-b677-40d3-b3a7-c68779533f7c.html", "High"],
    [REPORT_DATE, "Easley (Anderson/Pickens border)", "Anderson", "SC", "Signature Foods USA", "202", "$11.5M",
     "Estimated/Inferred sub-$50k (food mfg.)", "Operations confirmed online as of April 2026", "2026-04",
     "https://www.postandcourier.com/greenville/business/easley-anderson-county-signature-foods-investment/article_1d618a58-af96-435a-a05c-7cc6f9253a83.html", "High"],
    [REPORT_DATE, "Greenville", "Greenville", "SC", "GNQ Insilico — AI/quantum techbio HQ", "Not disclosed", "$500M (valuation, not capex)",
     "Not disclosed", "No update on headcount found this cycle", "2026-05",
     "https://www.greenvillebusinessmag.com/stories/techbio-innovator-gnq-moves-north-american-headquarters-labs-to-greenville,43132", "Med"],
    [REPORT_DATE, "North Charleston", "Charleston", "SC", "Boeing 787 — production-rate ramp", "1,000 (unchanged)", "$1.0B (existing)",
     "Estimated/Inferred mixed ($54-64k assemblers; eng. likely higher)", "Rate confirmed at 8/month as of Jan 2026 (up from 5/month a year prior); full 10/month target now more likely 2027 than fully sustained through 2026", "2026-01-27",
     "https://scdailygazette.com/2026/01/27/boeing-looks-to-build-on-momentum-at-sc-dreamliner-plant/", "Med-High"],
    [REPORT_DATE, "North Charleston", "Charleston", "SC", "SC Ports Authority — Leatherman Terminal Phase 2 (OPERATIONS PAUSED)", "400 direct + ~1,200 indirect (unchanged projection)", "$1.2B (terminal, existing)",
     "Not disclosed (Estimated mixed, some 6-figure specialist roles)", "REVERSAL: SC Ports announced 2026-06-25 it will pause ALL Leatherman Terminal operations starting Aug. 1, 2026, citing high ILA labor costs and weak container volume; cargo consolidates to Wando Welch/North Charleston terminals; Phase 2 wharf construction reportedly continues toward 2027 completion but no operations-resumption timeline given", "2026-06-25",
     "https://scdailygazette.com/2026/06/25/sc-ports-temporarily-halts-leatherman-operations-in-face-of-high-costs-waning-demand/", "High (risk)"],
    [REPORT_DATE, "Goose Creek", "Berkeley", "SC", "HII (Newport News Shipbuilding) — 1-yr anniversary", "~250 (estimate, not re-confirmed)", "Not disclosed",
     "Not disclosed", "One-year operations anniversary reached 2026-01-22; company reports exceeding production targets, but no updated job-count figure surfaced", "2026-01-22",
     "https://www.live5news.com/2026/01/22/defense-provider-celebrates-one-year-operations-lowcountry/", "Med"],
    [REPORT_DATE, "Liberty", "Randolph", "NC", "Toyota Battery Manufacturing NC (TBMNC) — EV/hybrid battery megaplant", "5,100 (proj. total); 3,000+ now employed", "$13.9B",
     "$62,234 avg (stated; below $100k)", "NEW to Watch: employment crossed 3,000+ workers (Feb 2026 milestone); all-EV battery line targeted to launch later 2026, plug-in hybrid line mid-2027; scale of hiring velocity in a rural county may pressure workforce housing despite sub-$100k wage", "2026-02",
     "https://www.wfmynews2.com/article/news/local/major-milestone-over-3000-people-now-work-at-the-toyota-battery-plant/83-b192b700-91b7-42f4-837e-33d0c01ee96e", "Med-High"],
    [REPORT_DATE, "Spartanburg (former Kohler plant)", "Spartanburg", "SC", "NorthMark Strategies \"Project Moc-1\" — AI/HPC data center", "~150 FT at full operation (revised up from ~27); 400-600 temporary construction", "$2.8B",
     "Not disclosed", "NEW to Watch: jobs figure revised upward; contested public hearing held 2026-06-25 (SCDES comment period through 2026-07-31); power-capacity expansion request (48MW->450MW) under review", "2026-06-25",
     "https://www.postandcourier.com/spartanburg/news/spartanburg-sc-data-center-northmark-power-tax/article_c3e02eef-fe42-4fda-9bd6-df50ec79a3ed.html", "Med (contested/evolving)"],
    [REPORT_DATE, "Greenville", "Greenville", "SC", "GE Vernova — gas turbine mfg (AI/data-center power supply chain)", "650 total planned; ~200 hired 2025, ~300 more targeted by end 2026", "$160M (Greenville); part of $600M multi-site plan",
     "Estimated/Inferred mixed — production/technician roles ~$45-56k; specialist/eng. roles $91k-$226k (Glassdoor); bulk of new hires likely sub-$100k", "NEW to Watch: in-window (2026-06-27) hiring-ramp confirmation; original Jan-2025 investment announcement was pre-window", "2026-06-27",
     "https://scbiz.com/ge-vernova-to-invest-160m-create-650-jobs-with-greenville-expansion/", "Med"],
    [REPORT_DATE, "RTP", "Wake", "NC", "Apple — RTP campus extension (incentive-timeline extension)", "3,000 originally projected; only ~600-990 added to date", "$552M-$1B+ (paused)",
     "$187k avg (stated, original 2021 award)", "NEW to Watch: NC Economic Investment Committee granted a four-year extension on hiring/investment timelines (April 2026) restarting the 2021 deal (up to $845M in potential tax benefits if targets met); no active construction found", "2026-04",
     "https://www.carolinajournal.com/apple-delays-construction-on-552-million-rtp-campus/", "Low-Med (long-delayed, no active construction)"],
]

WATCH = WEEK1_WATCH + WEEK2_WATCH + WEEK3_WATCH

# ============================================================================
# Excluded / Noise
# ============================================================================
EXCLUDED_HEADERS = ["report_date", "company_project", "market", "state", "jobs", "announce_date", "reason_excluded", "source_url"]

# ---- Week 1 (2026-06-24) ----
WEEK1_EXCLUDED = [
    ["2026-06-24", "Pacific Life", "Charlotte / South End", "NC", "301", "2025-10-28", "Pre-window (>6 mo back) + sub-500 (high wage $176k — re-flag if expanded)", "https://governor.nc.gov/news/press-releases/2025/10/28"],
    ["2026-06-24", "Novartis", "Durham/Morrisville", "NC", "700", "2025-11-19", "Pre-window (>6 mo back); June 2026 was groundbreaking follow-up", "https://www.commerce.nc.gov/news/press-releases/2025/11/19/novartis-expand-us-manufacturing-footprint-durham-and-wake-counties-adding-700-jobs-771-million"],
    ["2026-06-24", "Aspida Financial", "Durham", "NC", "1,000", "2025-11-19", "Pre-window (>6 mo back)", "https://www.commerce.nc.gov/news/press-releases/2025/11/19/financial-services-company-expand-durham-headquarters-1000-new-jobs"],
    ["2026-06-24", "Vulcan Elements", "Benson (Johnston)", "NC", "1,000", "2025-11-18", "Pre-window (>6 mo back)", "https://www.commerce.nc.gov/news/press-releases/2025/11/18/governor-stein-announces-vulcan-elements-selects-johnston-county-1000-job-magnet-factory-investing"],
    ["2026-06-24", "Maersk", "Charlotte", "NC", "520", "2025-11-18", "Pre-window (>6 mo back)", "https://www.sedc.org/"],
    ["2026-06-24", "Citigroup", "Charlotte", "NC", "~510", "2026-03-16", "In-window date but office grand opening tied to prior-year commitment, not a new 500+ award", "https://governor.nc.gov/news/press-releases/2026/03/16"],
    ["2026-06-24", "Jabil", "Salisbury (Rowan)", "NC", "1,181", "2025-06-30", "Pre-window (>6 mo back); wage $62k (< $100k)", "https://governor.nc.gov/news/press-releases/2025/06/30"],
    ["2026-06-24", "Boeing 787 site", "North Charleston", "SC", "1,000+", "2025-11-07", "Pre-window (>6 mo back, groundbreaking)", "https://www.prnewswire.com/news-releases/boeing-south-carolina-breaks-ground-on-787-site-expansion-302608798.html"],
    ["2026-06-24", "Google SC data centers", "Berkeley/Dorchester", "SC", "jobs n/d", "2025-10-13", "Pre-window (2025-10-13) + jobs undisclosed (also in Watch)", "https://blog.google/company-news/inside-google/company-announcements/google-american-innovation-south-carolina/"],
    ["2026-06-24", "Amazon robotics FC", "Pender Co.", "NC", "1,000+", "2025-03", "Pre-window + logistics wages (< $100k)", "https://businessnc.com/amazon-adding-1000-jobs-at-new-pender-county-center/"],
    ["2026-06-24", "Toyota Battery NC", "Liberty (Randolph)", "NC", "5,100 cum.", "2021-2023", "Out of window; ~$20.77/hr median", "https://abc11.com/post/toyota-nc-plant-investing-139-billion-creating-5100-jobs-randolph-county-north-carolina/18148348/"],
    ["2026-06-24", "Scout Motors (original selection)", "Blythewood", "SC", "4,000", "2023-03", "2023 selection — superseded here by the in-window hiring milestone (now ranked #2/#3)", "https://governor.sc.gov/news/2023-03/scout-motors-selects-south-carolina-production-site-plans-create-4000-jobs"],
    ["2026-06-24", "Goodyear plant", "Fayetteville (Cumberland)", "NC", "-", "2026-05", "LAYOFF/CLOSURE (excluded per rules)", "https://www.wral.com/business/fayetteville-cumberland-industrial-jobs-may-2026/"],
    ["2026-06-24", "Ralliant", "Raleigh (Wake)", "NC", "180", "2025-03-11", "Pre-window + sub-500 (high wage $189k)", "https://www.commerce.nc.gov/news/press-releases/2025/03/11/governor-stein-announces-180-jobs-global-technology-company-selects-wake-county-new-headquarters"],
    ["2026-06-24", "BuildOps", "Raleigh (Wake)", "NC", "291", "2025-06-24", "Pre-window + sub-500", "https://www.commerce.nc.gov/news/press-releases/2025/06/24/governor-stein-announces-software-company-buildops-will-create-291-jobs-raleigh"],
    ["2026-06-24", "Pallidus", "York Co.", "SC", "405", "2023-02", "Old project recycled", "https://www.yorkcountyed.com/news-media/announcements/pallidus-relocating-corporate-headquarters-and-manufacturing-operations-to-york-county"],
    ["2026-06-24", "E&J Gallo Winery", "Chester Co.", "SC", "~496", "2021-06", "Old project; recent items facility recognition", "https://governor.sc.gov/news/2021-06/e-j-gallo-winery-establishing-new-east-coast-facility-chester-county"],
    ["2026-06-24", "IKO North America", "Chester Co.", "SC", "~180", "2026-03-25", "Grand opening of 2023 project; sub-500", "https://charlotteregion.com/news/global-roofing-manufacturer-investing-363m-creating-180-jobs-in-chester-county/"],
    ["2026-06-24", "Apple / JPMorgan / BofA / Truist branches", "Charlotte", "NC", "n/d", "various", "No specific qualifying job count; branch-network/speculative", "n/a"],
]

# ---- Week 2 (2026-07-06) ----
WEEK2_EXCLUDED = [
    ["2026-07-06", "Wegmans", "Charlotte area (Ballantyne)", "NC", "450", "2026-01-15", "In-window but low-wage-only retail (~$14-16/hr)", "https://www.wbtv.com/2026/01/15/wegmans-bring-450-jobs-new-charlotte-area-store-how-apply/"],
    ["2026-07-06", "PSA Airlines", "Charlotte", "NC", "~400", "2026-03-19", "HQ grand opening of a prior Ohio-closure-driven relocation; mixed/mostly sub-$100k except pilots", "https://governor.nc.gov/news/press-releases/2026/03/19/governor-stein-celebrates-psa-airlines-headquarters-grand-opening-charlotte-highlights-only"],
    ["2026-07-06", "Maersk", "Charlotte", "NC", "520", "2025-11-18", "Pre-window; in-window 'milestone' is a 1-year hiring delay (2027-2029), not qualifying forward movement", "https://businessnc.com/maersk-delays-charlotte-headquarters-plan-hiring-520-workers-by-1-year/"],
    ["2026-07-06", "Pacific Life", "Charlotte / South End", "NC", "301", "2025-10-28", "Pre-window + sub-500; in-window milestone is only an interim-office opening", "https://www.pacificlife.com/press-releases/pacific-life-announces-footprint-expansion-to-charlotte-north-carolina.html"],
    ["2026-07-06", "Live Oak Bank", "Wilmington", "NC", "-", "2026-05-13", "Withdrew from state JDIG citing ~30-job hiring shortfall vs. 204-job target — negative reversal", "https://www.wilmingtonbiz.com/banking_and_finance/2026/05/13/live_oak_bank_pulls_out_of_state_incentive_program_citing_projected_hiring_shortfall/27496"],
    ["2026-07-06", "Technimark", "Asheboro (Randolph)", "NC", "-", "2026-02", "State terminated JDIG for missed hiring targets — negative reversal", "https://www.wral.com/news/nccapitol/north-carolina-terminates-incentives-six-companies-1500-jobs-feb-2026/"],
    ["2026-07-06", "Hoffman & Hoffman", "Greensboro (Guilford)", "NC", "131", "2025-12-24", "Announced 13 days before window opens", "https://www.areadevelopment.com/newsItems/12-24-2025/hoffman-hoffman-greensboro-north-carolina.shtml"],
    ["2026-07-06", "AssetMark Financial Holdings", "Charlotte", "NC", "252", "2025-07-08", "Pre-window; no in-window update despite high wage ($110,518)", "https://www.commerce.nc.gov/news/press-releases/2025/07/08/governor-stein-announces-financial-services-firm-assetmark-will-create-252-jobs-charlotte"],
    ["2026-07-06", "Fit Precast", "Gaston Co.", "NC", "125", "2025-11-18", "Pre-window + sub-200 jobs despite high wage ($104,000)", "https://businessnc.com/114420-2/"],
    ["2026-07-06", "Silfab Solar", "York Co.", "SC", "800", "2023-09", "Pre-window; no in-window milestone found", "https://pv-magazine-usa.com/"],
    ["2026-07-06", "Jabil", "Salisbury (Rowan)", "NC", "1,181", "2025-06-30", "Pre-window; wage $62,034 (<$100k)", "https://governor.nc.gov/news/press-releases/2025/06/30/jabil-selects-rowan-county-nearly-1200-new-jobs-and-500-million-multi-year-investment"],
    ["2026-07-06", "MetOx International", "Chatham Co.", "NC", "333", "2024-12-17", "Pre-window; wage $75,132 (<$100k)", "https://www.commerce.nc.gov/"],
    ["2026-07-06", "Amgen (2nd facility)", "Holly Springs (Wake)", "NC", "370", "2024-12", "Pre-window; wage >$91k but <$100k", "https://www.commerce.nc.gov/"],
    ["2026-07-06", "Biogen", "Durham/Wake (RTP)", "NC", "n/d", "2025-07-21", "Pre-window; no job figure disclosed", "https://www.biogen.com/"],
    ["2026-07-06", "Crystal Window & Door Systems", "Johnston Co.", "NC", "501", "2024-05-02", "Pre-window; meets job count but wage $56,061 (<$100k)", "https://www.commerce.nc.gov/"],
    ["2026-07-06", "Novo Nordisk (2nd facility)", "Clayton (Johnston)", "NC", "1,000", "2024-06-24", "Pre-window; wage ~$65-70k (<$100k)", "https://www.commerce.nc.gov/"],
    ["2026-07-06", "Apple RTP campus extension", "Wake Co.", "NC", "2,700+", "2025-11-25", "Pre-window; no in-window construction start confirmed", "https://www.apple.com/"],
    ["2026-07-06", "Pendo", "Raleigh (Wake)", "NC", "-90", "2026-04-07", "Layoff — excluded per rules", "https://www.pendo.io/"],
    ["2026-07-06", "\"Barclays 1,500 jobs, North Hills\" (unverified)", "Raleigh (Wake)", "NC", "1,500 (claimed)", "unverified", "Unverifiable — only source is a non-authoritative relocation/SEO blog; no corroboration found", "n/a"],
    ["2026-07-06", "\"Energizer Holdings HQ relocation to Apex\" (unverified)", "Apex (Wake)", "NC", "~75 (claimed)", "unverified", "Unverifiable — same non-authoritative source issue; Energizer's own newsroom shows no such announcement", "n/a"],
    ["2026-07-06", "GITI Tire", "Chester Co.", "SC", "1,700 cum.", "2015-2017", "Legacy project; no in-window milestone despite ongoing hiring", "n/a"],
    ["2026-07-06", "Eli Lilly", "Concord (Cabarrus)", "NC", "~600", "2022", "Legacy announcement/groundbreaking; wage >$70k (<$100k); no in-window update", "n/a"],
    ["2026-07-06", "Windsor Windows & Doors", "Monroe (Union)", "NC", "-", "2024", "Facility largely complete by 2024; no in-window figures disclosed", "n/a"],
    ["2026-07-06", "Digital Realty", "Mecklenburg Co.", "NC", "n/d", "2024-11", "Land acquired Nov. 2024; rezoning approved but zero jobs figure ever disclosed", "n/a"],
    ["2026-07-06", "Duke Energy gas plant", "Anderson Co.", "SC", "125 permanent", "2026-03-26", "Below the 200-job Watch floor despite a 2,200-job, multi-year construction surge (2027-2031)", "https://www.foxcarolina.com/"],
    ["2026-07-06", "GE Vernova / EnerSys / Woodward / Isuzu N.A. (legacy Upstate deals)", "Greenville/Spartanburg Cos.", "SC", "500-700 each", "2024-2025", "All pre-window; no confirmed 2026 groundbreaking/hiring-ramp milestone found", "n/a"],
    ["2026-07-06", "Sub-200-job Upstate SC items (Huwell, Coastal Precast, Hydrite, United Composite, DartPoints, DMA Industries, Carbotech, AFL/Fujikura, Project Moc-1, Advanced Metalworks)", "Various Upstate", "SC", "10-100 each", "2025-2026", "Below Watch floor regardless of window; Project Moc-1 has $2.76B capex but only 27 jobs", "n/a"],
    ["2026-07-06", "Sub-200-job Triad/Piedmont items (DePalo Foods, Textum OPCO, Cheerwine, ChenMed, Alamance Foods, Nature's Value, WH Farms, 7 Cinematics)", "Various Triad", "NC", "10-183 each", "2021-2026", "Below Watch floor and/or wage well under $100k", "n/a"],
]

# ---- Week 3 (2026-07-13) ----
WEEK3_EXCLUDED = [
    [REPORT_DATE, "Maersk North American HQ", "Charlotte", "NC", "520", "2025-11-18", "Pre-window; in-window milestone remains a 1-year hiring delay (2027-2029) confirmed again this cycle; wage now disclosed at $100,962 (just above $100k) but a delay is not qualifying forward movement per report rules", "https://businessnc.com/maersk-delays-charlotte-headquarters-plan-hiring-520-workers-by-1-year/"],
    [REPORT_DATE, "Citigroup", "Charlotte / Ballantyne", "NC", "510", "2025-07-08", "Pre-window; in-window milestone (3/16/26 grand opening) ties to a prior-year commitment, not a new 500+ award — consistent with this dataset's baseline-report exclusion rule", "https://hoodline.com/2026/03/stein-hits-ballantyne-as-citi-bets-big-on-charlotte-jobs/"],
    [REPORT_DATE, "TigerDC \"Project Spero\" (WITHDRAWN)", "Spartanburg Co.", "SC", "~50 FTE (Phase I)", "2026-02-27", "REVERSAL: Spartanburg County Council voted down the tax incentive; TigerDC formally withdrew the $3.0B project — removed from Watch tier this cycle", "https://www.foxcarolina.com/2026/02/27/company-withdraws-ai-data-center-spartanburg-county-consideration/"],
    [REPORT_DATE, "PSA Airlines", "Charlotte (CLT)", "NC", "~400-450 (mix reloc.+new; ~250 net new)", "2026-03-19", "No new in-window update; net new local roles (~250) remain well below 500-job threshold; wage mixed/mostly sub-$100k except pilots", "https://psaairlines.com/psa-airlines-celebrates-grand-opening-of-charlotte-headquarters-underscoring-economic-and-workforce-impact-across-north-carolina/"],
    [REPORT_DATE, "Scout Motors HQ (Plaza Midwood)", "Charlotte", "NC", "1,200", "2025-11-12", "Pre-window; no confirmed dated in-window milestone found (only vague 'spring/summer 2026 build-out' reporting) — recheck next cycle for a concrete groundbreaking/permit date", "https://www.detroitnews.com/story/business/autos/2025/11/14/scout-motors-rolls-into-charlotte-5-key-things-about-the-automaker/87270881007/"],
    [REPORT_DATE, "Bosch — Dorchester Co. electric motors", "Dorchester", "SC", "n/d (workforce 1,500->1,800)", "n/d", "Incremental headcount growth only; expansion physically paused due to EV slowdown; no discrete new-jobs announcement found", "https://www.postandcourier.com/business/bosch-dorchester-electric-vehicles-rivian-expansion/article_265b1c36-ab6c-11ef-b047-3b784368d210.html"],
    [REPORT_DATE, "Compass Datacenters", "Statesville (Iredell)", "NC", "~250", "2025-08", "Announcement and rezoning both pre-window; no new 2026 milestone found", "n/a"],
    [REPORT_DATE, "Southeastern Container", "Enka-Candler (Buncombe)", "NC", "12", "2026-02-19", "In-window but far below the 200-job Watch floor", "n/a"],
    [REPORT_DATE, "Amazon robotics fulfillment center", "Pender/New Hanover", "NC", "1,000", "2025-03", "Pre-window; fulfillment-center wages well under $100k, no offsetting wage signal found", "n/a"],
    [REPORT_DATE, "Walmart — Mebane distribution center", "Alamance", "NC", "450", "2025-09-30", "Pre-window; wage $35,374 (well under $100k)", "n/a"],
    [REPORT_DATE, "\"3 medical textile manufacturers\" (unnamed)", "Winston-Salem (Forsyth)", "NC", "600+ (aspirational)", "2026", "Speculative — no specific company named; aspirational recruitment target only, not a qualifying announcement", "n/a"],
    [REPORT_DATE, "\"Barclays 1,500 jobs, North Hills\" (unverified)", "Raleigh (Wake)", "NC", "1,500 (claimed)", "unverified", "Re-checked this cycle: still unverifiable — only source is a non-authoritative relocation-marketing content site; no corroboration from Barclays, WRAL, Triangle Business Journal, NC Commerce, or Governor's office", "n/a"],
    [REPORT_DATE, "Novo Nordisk speculative Four Oaks expansion", "Johnston", "NC", "~500 (speculative)", "2025-10 (cancelled)", "Project cancelled after an annexation vote in Oct. 2025; also pre-window", "n/a"],
    [REPORT_DATE, "Numerous sub-200-job items (Shamrock Technologies, Meiden America Switchgear, Coastal Precast, Hydrite Chemical, Huwell US, Milliken Cherokee, DartPoints, Advanced Metalworks, Eastern Engineered Wood, DMA Industries, United Composite, Aran USA, Carbotech, Peabody Engineering, ZF Chassis, Idealworks, Transcom)", "Various Upstate SC", "SC", "13-450 each", "2021-2026", "Below Watch floor regardless of window, or recycled old news (Transcom 2021, Bosch/Anderson 2022) with no in-window update", "n/a"],
    [REPORT_DATE, "Numerous sub-200-job items (SHL Medical, ATP Adhesives, Cardiff Products USA, The Nuclear Company, Carolina Renewable Products, AVM Group, Red Metals, Bittermilk Bottling, Modus21, Hoffman & Hoffman Lexington, Eastover Chips, Leonardo DRS, Nucor Berkeley galvanizing)", "Various Midlands/Lowcountry", "SC", "12-220 each", "2022-2026", "Below Watch floor and/or recycled old news with no in-window update", "n/a"],
]

EXCLUDED = WEEK1_EXCLUDED + WEEK2_EXCLUDED + WEEK3_EXCLUDED

# ============================================================================
# Source Log
# ============================================================================
SOURCE_HEADERS = ["source_name", "tier", "url", "coverage", "used_for"]
SOURCES = [
    ["NC Commerce — Press Releases", "state_primary", "https://www.commerce.nc.gov/news/press-releases", "NC statewide", "SMBC, AbbVie, JetZero, BorgWarner (stated wages + JDIG)"],
    ["NC Governor's Office", "state_primary", "https://governor.nc.gov/news/press-releases", "NC statewide", "SMBC, Capital Group, JetZero, PSA Airlines"],
    ["EDPNC — News", "state_primary", "https://edpnc.com/news/", "NC statewide", "Cross-check"],
    ["SC Commerce — News", "state_primary", "https://www.sccommerce.com/news", "SC statewide", "USA Rare Earth, FN America, AESC"],
    ["SC Governor's Office", "state_primary", "https://governor.sc.gov/news", "SC statewide", "Ferrara, Suniva, AMAROK, Scout (project), Octapharma context"],
    ["AbbVie press release", "company_primary", "https://news.abbvie.com/", "Company", "AbbVie campus"],
    ["Genentech press release", "company_primary", "https://www.gene.com/media/press-releases", "Company", "Genentech Holly Springs investment doubling (Jan 2026)"],
    ["Scout Motors company blog", "company_primary", "https://blog.scoutmotors.com/", "Company", "June 2026 production-center update (first vehicle body weld)"],
    ["Trade & Industry Dev / Carolina Journal", "journal", "https://www.tradeandindustrydev.com/", "NC/SC", "Siemens Energy $421M/500-job NC expansion; JetZero hiring delay + $133.9M funding"],
    ["WRAL / WRAL TechWire", "journal", "https://www.wral.com/", "NC Triangle", "AbbVie salary/incentive detail; Novartis $220M expansion; Siemens AG"],
    ["WIS / WLTX (Columbia)", "journal", "https://www.wistv.com/", "SC Midlands", "Scout Motors hiring milestone, wages; AMAROK"],
    ["Charlotte Business Journal / Axios Charlotte", "journal", "https://charlotte.axios.com/", "Charlotte metro", "Capital Group corroboration; SMBC lease; JPMorgan Chase; Pacific Life"],
    ["Area Development", "journal", "https://www.areadevelopment.com/", "National site-selection", "Lumentum; Hoffman & Hoffman"],
    ["Post & Courier", "journal", "https://www.postandcourier.com/", "SC statewide", "QTS York County; Octapharma Rock Hill; SC Ports; AMAROK"],
    ["Asheville Chamber", "regional_edo", "https://www.ashevillechamber.org/", "Buncombe NC", "Eaton expansion; post-Helene recovery context"],
    ["Charlotte Regional Business Alliance", "regional_edo", "https://charlotteregion.com/", "CLT 15-county bi-state", "IKO/Chester cross-check"],
    ["York County Economic Development", "regional_edo", "https://www.yorkcountyed.com/", "York Co. SC", "Pallidus (excluded); Octapharma context"],
    ["Upstate SC Alliance / Fox Carolina", "regional_edo/journal", "https://www.upstatescalliance.com/", "Upstate SC", "Siemens Smart Infra; TigerDC Project Spero"],
    ["Business NC", "journal", "https://businessnc.com/", "NC statewide", "Siemens Energy; Pratt & Whitney; Maersk delay; Capital Group lease"],
    ["Fort Mill Sun / WRHI / WSOC-TV / QC News", "journal", "https://www.fortmillsun.com/", "Charlotte-metro SC (York Co.)", "Octapharma 'Project Palmetto Rock' council vote + controversy"],
    ["Commercial Real Estate Direct", "journal", "https://crenews.com/", "National CRE", "Capital Group lease at One Independence Center"],
    ["Salisbury Post", "journal", "https://www.salisburypost.com/", "Rowan Co. NC", "Project Rack; Google/DHL Kannapolis lease"],
    ["Hoodline", "journal", "https://hoodline.com/", "NC (Triad/Triangle)", "Boom Supersonic risk flag; Ralliant HQ launch"],
    ["Chatham County EDC", "regional_edo", "https://www.chathamedc.org/", "Chatham Co. NC", "Wolfspeed Siler City milestone"],
    ["WUNC", "journal", "https://www.wunc.org/", "NC Triangle", "Vulcan Elements training milestone; WakeMed/Atrium merger"],
    ["ABC11 / NC Health News", "journal", "https://abc11.com/", "NC Triangle", "WakeMed/Atrium Health merger proposal"],
    ["Aviation Week / Manufacturing Dive", "trade_journal", "https://aviationweek.com/", "National aerospace", "Boeing 787 production-rate ramp"],
    ["WilmingtonBiz", "journal", "https://www.wilmingtonbiz.com/", "New Hanover Co. NC", "Amazon-Corning fiber optics; Live Oak Bank JDIG withdrawal"],
    ["SC Daily Gazette", "journal", "https://scdailygazette.com/", "SC statewide", "USA Rare Earth wage detail; Leatherman Terminal pause; Boeing rate-ramp"],
    ["Live5News / Berkeley Observer", "journal", "https://www.live5news.com/", "Charleston/Lowcountry SC", "HII one-year anniversary; SC Ports Leatherman pause"],
    ["SEC EDGAR filings", "primary_filing", "https://www.sec.gov/", "National", "Boeing 787 10-Q rate-ramp confirmation; Wolfspeed 10-Q post-bankruptcy status"],
    ["Post and Courier — Spartanburg/Greenville bureaus", "journal", "https://www.postandcourier.com/spartanburg/", "Upstate SC", "TigerDC withdrawal; NorthMark Project Moc-1; AIRSYS timeline slip; GE Vernova; Suniva"],
    ["World Nuclear News", "trade_journal", "https://www.world-nuclear-news.org/", "National nuclear", "GE Hitachi GNF4 fuel-line fabrication start"],
    ["WFMY News 2", "journal", "https://www.wfmynews2.com/", "NC Triad", "Toyota Battery Manufacturing NC 3,000-employee milestone"],
    ["Corning press releases", "company_primary", "https://www.corning.com/", "Company", "Amazon-Corning fiber-optics partnership wage/job detail"],
    ["WBTV / Fort Mill Sun (Charlotte-metro follow-up)", "journal", "https://www.wbtv.com/", "Charlotte metro & York Co. SC", "Octapharma 3rd-reading vote; Maersk/Citigroup verification"],
]

# ============================================================================
# Scoring Methodology
# ============================================================================
SCORING_HEADERS = ["factor", "weight", "what_it_rewards"]
SCORING = [
    ["Job count", 25, "Absolute headcount (scaled; 2,000+ ~ full, 500 ~ ~12)"],
    ["Salary / income quality", 25, "Stated avg wage vs $100k bar; inferred wages discounted"],
    ["Employer & industry strength", 15, "Balance-sheet/brand durability, industry growth, execution certainty"],
    ["Capital investment / project certainty", 15, "Capex size + lock-in (incentive-backed, under construction)"],
    ["MF / townhome demand relevance", 15, "Urban/inner-suburban fit; income tier match to Class-A / for-sale TH"],
    ["Repeat momentum", 5, "Reinforcement of a market across reports (new = low; gaining-momentum = higher)"],
    ["TOTAL", 100, ""],
]

SCORE_DETAIL_HEADERS = ["report_date", "company_project", "score", "score_rationale"]
WEEK1_SCORE_DETAIL = [
    ["2026-06-24", "SMBC Group (Charlotte)", 92, "2,000 jobs + stated $165k + top-tier employer + urban Class-A fit; new (low momentum)."],
    ["2026-06-24", "Scout Motors (Blythewood)", 85, "4,000 jobs + $2B locked-in + active hiring; wage mixed/workforce-to-salaried."],
    ["2026-06-24", "AbbVie (Durham)", 84, "734 jobs + stated $118k + $1.4B top-pharma certainty; strong MF fit."],
    ["2026-06-24", "JetZero (Greensboro)", 82, "14,500 jobs (max) + $4.7B; wage $89k (<$100k) + budget/timeline execution risk discount."],
    ["2026-06-24", "Genentech (Holly Springs)", 80, "Site crosses 500 jobs + stated ~$120k + ~$2B (doubled) top-pharma certainty; in-window increment is +100, slight discount."],
    ["2026-06-24", "Capital Group (Charlotte)", 78, "600 jobs + inferred ~$190k (discounted vs stated) + urban fit; new."],
]
WEEK2_SCORE_DETAIL = [
    ["2026-07-06", "SMBC Group (Charlotte)", 93, "Unchanged fundamentals + lease/hiring milestone reinforces momentum (+1)."],
    ["2026-07-06", "Octapharma (Rock Hill)", 88, "1,500+ jobs + two stated six-figure wage tiers (rare for SC) + $1.5B; discounted for tax-share controversy/execution risk."],
    ["2026-07-06", "Scout Motors (Blythewood)", 86, "First vehicle body welded is a lower-risk production milestone than training-center opening alone (+1)."],
    ["2026-07-06", "AbbVie (Durham)", 84, "Unchanged; no new milestone this cycle."],
    ["2026-07-06", "JetZero (Greensboro)", 81, "State funding secured (+) but company's own hiring deadline slipped a full year (-); net -1."],
    ["2026-07-06", "Genentech (Holly Springs)", 80, "Unchanged; no new milestone this cycle."],
    ["2026-07-06", "Capital Group (Charlotte)", 79, "Uptown lease signed reinforces momentum (+1)."],
    ["2026-07-06", "Novartis (Durham/Morrisville)", 77, "700 jobs + stated $111,161 + top-pharma certainty; newly in-window via investment-increase milestone, low repeat-momentum score as first-time-ranked."],
]
WEEK3_SCORE_DETAIL = [
    [REPORT_DATE, "SMBC Group (Charlotte)", 93, "Unchanged; no new milestone this cycle."],
    [REPORT_DATE, "Octapharma (Rock Hill)", 87, "County-level 3rd reading approved but on a narrower 4-3 vote; Rock Hill City Council's vote on the amended tax-share terms remains outstanding — net -1 for continued execution uncertainty despite forward progress."],
    [REPORT_DATE, "Scout Motors (Blythewood)", 86, "Active hiring push continues (new July 29 interview events, ~1,400 hired companywide); unchanged despite a mild timeline slip in customer deliveries."],
    [REPORT_DATE, "AbbVie (Durham)", 84, "Unchanged; no new milestone this cycle."],
    [REPORT_DATE, "Capital Group (Charlotte)", 81, "Wage figure upgraded from Estimated/Inferred to a stated $194,141 avg (NC Commerce JDIG record); +2 for removed wage-confidence discount."],
    [REPORT_DATE, "Genentech (Holly Springs)", 80, "Unchanged; no new milestone this cycle."],
    [REPORT_DATE, "JetZero (Greensboro)", 79, "Groundbreaking held (tangible positive) but JDIG amendment confirms hiring ramp pushed out further (full target now 2037 vs. 2036); net -2."],
    [REPORT_DATE, "Novartis (Durham/Morrisville)", 77, "Unchanged; no new milestone this cycle."],
]
SCORE_DETAIL = WEEK1_SCORE_DETAIL + WEEK2_SCORE_DETAIL + WEEK3_SCORE_DETAIL

# ============================================================================
# Weekly Summary
# ============================================================================
SUMMARY_LINES = [
    ("Report", "Carolinas Job Growth & Housing Demand Report"),
    ("Week / Report date", REPORT_DATE + "  (Week 3)"),
    ("Coverage window", WINDOW + "  (trailing 6 months)"),
    ("Geography", "North Carolina + South Carolina"),
    ("Qualifying (ranked) deals this week", str(len(WEEK3_RUNNING))),
    ("Qualifying (ranked) deals — cumulative", str(len(RUNNING))),
    ("Markets to Watch this week", str(len(WEEK3_WATCH))),
    ("Excluded / Noise this week", str(len(WEEK3_EXCLUDED))),
    ("", ""),
    ("Top market", "Charlotte / Uptown (Mecklenburg, NC) — SMBC Group, score 93"),
    ("Top mover", "Rock Hill / York Co. (SC) — Octapharma 'Project Palmetto Rock', score 87 — county-level 3rd reading approved 4-3 but Rock Hill City Council's vote on amended tax-share terms remains outstanding"),
    ("Key takeaway", "No brand-new qualifying (500+ job / $100k+ wage) deal was found in any of the five sub-regions this cycle — all 8 ranked entries carry forward. JetZero's groundbreaking is real construction progress even as its hiring ramp slips further (full target now 2037). SC Ports' decision to pause Leatherman Terminal operations Aug. 1 and Boom Supersonic's looming Dec. 31 hiring deadline are the cycle's clearest downside signals; TigerDC's $3.0B Spartanburg data center was formally withdrawn."),
    ("Window note", "Trailing 6 months now runs 2026-01-13 to 2026-07-13."),
    ("Data-quality note", "Two near-miss candidates (Maersk — wage now disclosed at $100,962 but milestone is a hiring delay; Citigroup — grand opening tied to a pre-window commitment) were excluded for consistency with this dataset's established milestone rules rather than admitted on job count/wage alone."),
    ("Week-over-week", "8 of 8 prior-ranked deals remain qualifying (none removed, none newly added); 1 gaining momentum (Scout Motors), 2 updated-mixed (Octapharma, JetZero), 1 updated (Capital Group, wage confirmation), 4 repeated (SMBC, AbbVie, Genentech, Novartis)."),
]


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def write_table(ws, headers, rows, start_row=1):
    for j, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=j, value=h)
    style_header(ws, start_row, len(headers))
    for i, row in enumerate(rows, start_row + 1):
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=i, column=j, value=val)
            cell.alignment = WRAP
            cell.border = BORDER
            cell.font = Font(size=9)
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def set_widths(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


wb = Workbook()

# 1. Weekly Summary
ws = wb.active
ws.title = "Weekly Summary"
ws["A1"] = "Carolinas Job Growth & Housing Demand Report"
ws["A1"].font = TITLE_FONT
ws["A2"] = f"Week of {REPORT_DATE}  ·  Window {WINDOW}  ·  NC + SC"
ws["A2"].font = SUB_FONT
r = 4
for k, v in SUMMARY_LINES:
    ws.cell(row=r, column=1, value=k).font = Font(bold=True, size=10)
    c = ws.cell(row=r, column=2, value=v)
    c.alignment = WRAP
    c.font = Font(size=10)
    r += 1
set_widths(ws, {"A": 26, "B": 95})

# 2. Ranked Opportunities (subset of Running Database, all weeks to date)
ws = wb.create_sheet("Ranked Opportunities")
RANKED_HEADERS = ["report_date", "rank", "score", "market_submarket", "county", "state", "company_project",
                  "job_count", "job_type", "salary_wage_stated", "salary_inferred",
                  "capital_investment", "incentives", "announcement_type", "source_date",
                  "source_url", "confidence", "urban_suburban_rural", "housing_demand_implication"]
ranked_rows = []
idx = {h: RUNNING_HEADERS.index(h) for h in RANKED_HEADERS}
for row in RUNNING:
    ranked_rows.append([row[idx[h]] for h in RANKED_HEADERS])
write_table(ws, RANKED_HEADERS, ranked_rows)
set_widths(ws, {"A": 12, "B": 6, "C": 7, "D": 20, "E": 12, "F": 6, "G": 30, "H": 12, "I": 18,
                "J": 22, "K": 26, "L": 12, "M": 28, "N": 22, "O": 13, "P": 40, "Q": 16,
                "R": 14, "S": 50})

# 3. Running Database (full schema, one row per announcement, all weeks — ACCUMULATES)
ws = wb.create_sheet("Running Database")
write_table(ws, RUNNING_HEADERS, RUNNING)
widths = {get_column_letter(i): 18 for i in range(1, len(RUNNING_HEADERS) + 1)}
widths.update({"A": 12, "B": 22, "C": 5, "D": 6, "E": 20, "H": 30, "S": 40, "W": 45, "X": 40, "Q": 26, "P": 28})
set_widths(ws, widths)

# 4. Markets to Watch (all weeks — accumulates)
ws = wb.create_sheet("Markets to Watch")
write_table(ws, WATCH_HEADERS, WATCH)
set_widths(ws, {"A": 12, "B": 24, "C": 14, "D": 6, "E": 34, "F": 16, "G": 14,
                "H": 34, "I": 44, "J": 12, "K": 40, "L": 16})

# 5. Excluded / Noise (all weeks — accumulates)
ws = wb.create_sheet("Excluded_Noise")
write_table(ws, EXCLUDED_HEADERS, EXCLUDED)
set_widths(ws, {"A": 12, "B": 32, "C": 22, "D": 6, "E": 10, "F": 12, "G": 46, "H": 44})

# 6. Source Log
ws = wb.create_sheet("Source Log")
write_table(ws, SOURCE_HEADERS, SOURCES)
set_widths(ws, {"A": 34, "B": 20, "C": 46, "D": 24, "E": 44})

# 7. Scoring Methodology
ws = wb.create_sheet("Scoring Methodology")
write_table(ws, SCORING_HEADERS, SCORING)
set_widths(ws, {"A": 36, "B": 10, "C": 60})
start = len(SCORING) + 3
ws.cell(row=start, column=1, value="Per-deal score rationale (all weeks)").font = Font(bold=True, size=11, color="1F3864")
write_table(ws, SCORE_DETAIL_HEADERS, SCORE_DETAIL, start_row=start + 1)
ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 32
ws.column_dimensions["C"].width = 10
ws.column_dimensions["D"].width = 75

out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Carolinas_Job_Growth_Running_Database.xlsx")
wb.save(out)
print("Saved", out)
print("Tabs:", wb.sheetnames)
print("Ranked deals (cumulative):", len(RUNNING), "| Watch (cumulative):", len(WATCH), "| Excluded (cumulative):", len(EXCLUDED))
